from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, make_response
import os, io, zipfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pdfkit
from models import (
    Invoice,
    InvoiceItem,
    CreditNote,
    CreditNoteItem,
    Party,
    Company,
    db,
)
from forms import InvoiceForm, CreditNoteForm
from utils import login_required, parse_date, validate_tax_rates, number_to_words, get_current_company, get_export_path
from decimal import Decimal, ROUND_HALF_UP



invoices_bp = Blueprint('invoices', __name__)

@invoices_bp.route("/invoices", methods=["GET"])
@invoices_bp.route("/invoices/", methods=["GET"])
@login_required
def manage_invoices():
    year = request.args.get("year")
    month = request.args.get("month")
    # Only set defaults if not explicitly provided (for first load)
    if year is None:
        year = str(datetime.now().year)
    if month is None:
        month = str(datetime.now().month).zfill(2)
    search = request.args.get("search", "").strip()
    sort_by = request.args.get("sort_by", "date")
    sort_dir = request.args.get("sort_dir", "desc")
    selected_year = year
    selected_month = month
    search_query = search
    tax_type_filter = request.args.get("tax_type", "")
    party_filter = request.args.get("party", "")
    status_filter = request.args.get("status", "")

    query = Invoice.query.join(Party)

    if year:
        query = query.filter(db.extract("year", Invoice.invoice_date) == int(year))
    if month:
        query = query.filter(db.extract("month", Invoice.invoice_date) == int(month))
    if search:
        query = query.filter(
            db.or_(
                Invoice.invoice_no.ilike(f"%{search}%"),
                Invoice.reference_serial_no.ilike(f"%{search}%"),
                Party.gstin.ilike(f"%{search}%"),
                Party.name.ilike(f"%{search}%"),
            )
        )
    if tax_type_filter:
        query = query.filter(Invoice.tax_type == tax_type_filter)
    if party_filter:
        query = query.filter(Invoice.party_id == int(party_filter))
    if status_filter == "pending":
        query = query.filter(Invoice.invoice_no.is_(None))
    elif status_filter == "completed":
        query = query.filter(Invoice.invoice_no != None)

    sort_column = Invoice.invoice_no
    if sort_by == "date":
        sort_column = Invoice.invoice_date
    elif sort_by == "party":
        sort_column = Party.name

    if sort_dir == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    invoices_list = query.all()
    parties_list = Party.query.all()
    pending_count = Invoice.query.filter(Invoice.invoice_no.is_(None)).count()

    all_invoices = Invoice.query.all()
    years = sorted(
        set(inv.invoice_date.year for inv in all_invoices if inv.invoice_date)
    )
    if not years:
        years = [datetime.now().year]

    return render_template(
        "invoice_management.html",
        invoices=invoices_list,
        parties=parties_list,
        selected_year=selected_year,
        selected_month=selected_month,
        search_query=search_query,
        years=years,
        sort_by=sort_by,
        sort_dir=sort_dir,
        tax_type_filter=tax_type_filter,
        party_filter=party_filter,
        status_filter=status_filter,
        pending_count=pending_count,
    )


@invoices_bp.route("/invoice/edit/<int:invoice_id>", methods=["GET", "POST"])
@login_required
def edit_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    
    if request.method == "POST":
        invoice.invoice_no = request.form.get("invoice_no", "").strip() or None
        invoice.invoice_date = datetime.strptime(request.form.get("invoice_date"), "%Y-%m-%d").date()
        invoice.tax_type = request.form.get("tax_type")
        invoice.place_of_supply = request.form.get("place_of_supply")
        invoice.reverse_charge = float(request.form.get("reverse_charge") or 0)
        invoice.is_rcm = request.form.get("is_rcm") == "on"
        invoice.distributor_code = request.form.get("distributor_code", "").strip()
        
        party_id = request.form.get("party_id")
        if party_id:
            party = db.session.get(Party, int(party_id))
            if party:
                invoice.party_name = party.name
                invoice.party_address = party.address
                invoice.party_gstin = party.gstin
                invoice.party_pan = party.pan
                invoice.party_state = party.state
                invoice.party_state_code = party.state_code
        
        invoice.company_name = request.form.get("company_name", "")
        invoice.company_address = request.form.get("company_address", "")
        invoice.company_gstin = request.form.get("company_gstin", "")
        invoice.company_pan = request.form.get("company_pan", "")
        
        # Delete existing items
        for item in invoice.items:
            db.session.delete(item)
        
        # Add new items
        items_desc = request.form.getlist("item_description[]")
        items_taxable = request.form.getlist("item_taxable_value[]")
        items_cgst_rate = request.form.getlist("item_cgst_rate[]")
        items_sgst_rate = request.form.getlist("item_sgst_rate[]")
        items_igst_rate = request.form.getlist("item_igst_rate[]")
        
        for i in range(len(items_desc)):
            if items_desc[i].strip():
                taxable = float(items_taxable[i]) if i < len(items_taxable) else 0
                cgst_rate = float(items_cgst_rate[i]) if i < len(items_cgst_rate) else 0
                sgst_rate = float(items_sgst_rate[i]) if i < len(items_sgst_rate) else 0
                igst_rate = float(items_igst_rate[i]) if i < len(items_igst_rate) else 0
                
                taxable_dec = Decimal(str(taxable))
                if invoice.tax_type == "INTRA":
                    cgst_amt = float((taxable_dec * Decimal(str(cgst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                    sgst_amt = float((taxable_dec * Decimal(str(sgst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                    igst_amt = 0.0
                elif invoice.tax_type == "INTER":
                    cgst_amt = 0.0
                    sgst_amt = 0.0
                    igst_amt = float((taxable_dec * Decimal(str(igst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                else:
                    cgst_amt = sgst_amt = igst_amt = 0.0
                
                item = InvoiceItem(
                    invoice_id=invoice.id,
                    description=items_desc[i],
                    taxable_value=taxable,
                    cgst_rate=cgst_rate,
                    cgst_amt=cgst_amt,
                    sgst_rate=sgst_rate,
                    sgst_amt=sgst_amt,
                    igst_rate=igst_rate,
                    igst_amt=igst_amt,
                )
                db.session.add(item)
        
        invoice.total_in_words = number_to_words(invoice.calculate_gst()['total'])
        
        db.session.commit()
        flash("Invoice updated successfully", "success")
        return redirect(url_for("invoices.manage_invoices"))
    
    parties_list = Party.query.order_by(Party.name).all()
    companies_list = Company.query.order_by(Company.name).all()
    return render_template("edit_invoice.html", invoice=invoice, parties=parties_list, companies=companies_list)


@invoices_bp.route("/invoice/create", methods=["GET", "POST"])
def create_invoice():
    parties_list = Party.query.order_by(Party.name).all()
    
    if request.method == "POST":
        party_id = request.form.get("party_id")
        invoice_no = request.form.get("invoice_no", "").strip() or None
        reference_serial_no = request.form.get("reference_serial_no", "").strip()

        if not party_id:
            flash("Party selection is required.", "danger")
            return redirect(url_for("invoices.create_invoice"))

        if reference_serial_no:
            existing = Invoice.query.filter(
                Invoice.reference_serial_no == reference_serial_no
            ).first()
            if existing:
                flash(
                    f"Reference serial number {reference_serial_no} already exists",
                    "danger",
                )
                return redirect(url_for("invoices.create_invoice"))

        invoice_date = datetime.strptime(
            request.form.get("invoice_date"), "%Y-%m-%d"
        ).date()
        tax_type = request.form.get("tax_type")
        place_of_supply = request.form.get("place_of_supply")
        reverse_charge = float(request.form.get("reverse_charge") or 0)
        is_rcm = request.form.get("is_rcm") == "on"
        distributor_code = request.form.get("distributor_code", "").strip()

        if is_rcm and reverse_charge <= 0:
            flash("Reverse Charge amount is required when RCM is enabled.", "danger")
            return redirect(url_for("invoices.create_invoice"))

        invoice = Invoice(
            invoice_no=invoice_no,
            reference_serial_no=reference_serial_no,
            invoice_date=invoice_date,
            party_id=party_id,
            tax_type=tax_type,
            place_of_supply=place_of_supply,
            reverse_charge=reverse_charge,
            is_rcm=is_rcm,
            distributor_code=distributor_code,
        )

        party = db.session.get(Party, party_id)
        if party:
            invoice.party_name = party.name
            invoice.party_address = party.address
            invoice.party_gstin = party.gstin
            invoice.party_pan = party.pan
            invoice.party_state = party.state
            invoice.party_state_code = party.state_code

        invoice.company_name = request.form.get("company_name", "")
        invoice.company_address = request.form.get("company_address", "")
        invoice.company_gstin = request.form.get("company_gstin", "")
        invoice.company_pan = request.form.get("company_pan", "")

        db.session.add(invoice)
        db.session.flush()

        items_desc = request.form.getlist("item_description[]")
        items_sac_hsn = request.form.getlist("item_sac_hsn[]")
        items_taxable = request.form.getlist("item_taxable_value[]")
        items_cgst_rate = request.form.getlist("item_cgst_rate[]")
        items_sgst_rate = request.form.getlist("item_sgst_rate[]")
        items_igst_rate = request.form.getlist("item_igst_rate[]")

        for i in range(len(items_desc)):
            if items_desc[i].strip():
                taxable = float(items_taxable[i]) if i < len(items_taxable) else 0
                cgst_rate = float(items_cgst_rate[i]) if i < len(items_cgst_rate) else 0
                sgst_rate = float(items_sgst_rate[i]) if i < len(items_sgst_rate) else 0
                igst_rate = float(items_igst_rate[i]) if i < len(items_igst_rate) else 0
                sac_hsn = items_sac_hsn[i] if i < len(items_sac_hsn) else None

                is_valid, error_msg = validate_tax_rates(
                    tax_type, is_rcm, cgst_rate, sgst_rate, igst_rate
                )
                if not is_valid:
                    flash(f"Item '{items_desc[i][:30]}': {error_msg}", "danger")
                    return redirect(url_for("invoices.create_invoice"))

                taxable_dec = Decimal(str(taxable))
                if tax_type == "INTRA":
                    cgst_amt = float((taxable_dec * Decimal(str(cgst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                    sgst_amt = float((taxable_dec * Decimal(str(sgst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                    igst_amt = 0.0
                elif tax_type == "INTER":
                    cgst_amt = 0.0
                    sgst_amt = 0.0
                    igst_amt = float((taxable_dec * Decimal(str(igst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                else:
                    cgst_amt = sgst_amt = igst_amt = 0.0

                item = InvoiceItem(
                    invoice_id=invoice.id,
                    description=items_desc[i],
                    sac_hsn_code=sac_hsn,
                    taxable_value=taxable,
                    cgst_rate=cgst_rate,
                    cgst_amt=cgst_amt,
                    sgst_rate=sgst_rate,
                    sgst_amt=sgst_amt,
                    igst_rate=igst_rate,
                    igst_amt=igst_amt,
                )
                db.session.add(item)

        invoice.total_in_words = number_to_words(invoice.calculate_gst()['total'])

        db.session.commit()
        flash("Invoice created successfully", "success")
        return redirect(url_for("invoices.manage_invoices"))

    company = get_current_company()
    companies_list = Company.query.order_by(Company.name).all()
    return render_template("create_invoice.html", company=company, companies=companies_list, parties=parties_list)


@invoices_bp.route("/invoice/delete/<int:invoice_id>")
@login_required
def delete_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    db.session.delete(invoice)
    db.session.commit()
    flash("Invoice deleted successfully", "success")
    return redirect(url_for("dashboard.dashboard"))


@invoices_bp.route("/exports/view/<path:filename>")
@login_required
def view_export(filename):
    from flask import current_app
    file_path = os.path.join(current_app.config["EXPORT_FOLDER"], filename)
    if not os.path.exists(file_path):
        return "File not found", 404
    
    if filename.endswith(".pdf"):
        mimetype = "application/pdf"
    elif filename.endswith(".xlsx"):
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif filename.endswith(".zip"):
        mimetype = "application/zip"
    else:
        mimetype = "application/octet-stream"
        
    return send_file(file_path, as_attachment=False, mimetype=mimetype)


@invoices_bp.route("/invoice/view/<int:invoice_id>")
@login_required
def view_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    settings = {
        "company_name": invoice.company_name or "Not set",
        "address": invoice.company_address or "",
        "gstin": invoice.company_gstin or "",
        "pan": invoice.company_pan or "",
        "logo": "",
        "place_of_supply": invoice.place_of_supply or "",
        "state_code": "",
    }
    return render_template("invoice_preview.html", invoice=invoice, settings=settings)


@invoices_bp.route("/invoice/pdf/<int:invoice_id>")
@login_required
def generate_pdf(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    settings = {
        "company_name": invoice.company_name or "Not set",
        "address": invoice.company_address or "",
        "gstin": invoice.company_gstin or "",
        "pan": invoice.company_pan or "",
        "logo": "",
        "place_of_supply": invoice.place_of_supply or "",
        "state_code": "",
    }
    html_content = render_template("invoice_pdf_template.html", invoice=invoice, settings=settings)
    pdf_path = get_export_path(invoice)
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    pdfkit.from_string(html_content, pdf_path)
    return send_file(pdf_path, as_attachment=False, mimetype='application/pdf')


@invoices_bp.route("/invoice/preview-html/<int:invoice_id>")
@login_required
def preview_html(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    return render_template("invoice_print.html", invoice=invoice)


@invoices_bp.route("/invoice/generate-numbers", methods=["POST"])
@login_required
def generate_numbers():
    from utils import generate_invoice_numbers
    count = generate_invoice_numbers()
    if count > 0:
        flash(f"Generated invoice numbers for {count} invoices", "success")
    else:
        flash("No pending invoices to generate numbers for", "warning")
    return redirect(url_for("invoices.manage_invoices"))


@invoices_bp.route("/batch/export", methods=["POST"])
@login_required
def batch_export():
    from flask import current_app
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = [x.strip() for x in ids_raw.split(",") if x.strip()]

    if not invoice_ids:
        flash("No invoices selected", "warning")
        return redirect(url_for("dashboard.dashboard"))

    zip_path = os.path.join(
        current_app.config["EXPORT_FOLDER"],
        f"invoices_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
    )

    exported_count = 0
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for invoice_id in invoice_ids:
            invoice = db.session.get(Invoice, int(invoice_id))
            if not invoice or not invoice.invoice_no or not invoice.locked:
                continue

            settings = {
                "company_name": invoice.company_name if invoice.company_name else "Not set",
                "address": invoice.company_address or "",
                "gstin": invoice.company_gstin or "",
                "pan": invoice.company_pan or "",
                "logo": "",
                "place_of_supply": invoice.place_of_supply or "",
                "state_code": "",
            }
            html_content = render_template("invoice_pdf_template.html", invoice=invoice, settings=settings)
            pdf_path = get_export_path(invoice)
            os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
            pdfkit.from_string(html_content, pdf_path)
            zip_file.write(pdf_path, os.path.basename(pdf_path))
            os.remove(pdf_path)
            exported_count += 1

    if exported_count == 0:
        flash("No locked invoices with invoice numbers to export", "warning")
        return redirect(url_for("invoices.manage_invoices"))

    flash(f"File: {os.path.basename(zip_path)} saved in exports folder", "success")
    return redirect(url_for("invoices.manage_invoices"))


@invoices_bp.route("/invoices/batch-lock", methods=["POST"])
@login_required
def batch_lock():
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = [int(x) for x in ids_raw.split(",") if x.strip()] if ids_raw else []
    
    if not invoice_ids:
        flash("No invoices selected", "warning")
        return redirect(url_for("invoices.manage_invoices"))
    
    newly_locked = 0
    skipped_missing_no = 0
    already_locked = 0
    
    for invoice_id in invoice_ids:
        invoice = db.session.get(Invoice, invoice_id)
        if invoice:
            if invoice.invoice_no:
                if invoice.locked:
                    already_locked += 1
                else:
                    invoice.locked = True
                    newly_locked += 1
            else:
                skipped_missing_no += 1
    
    db.session.commit()
    
    if newly_locked > 0 and already_locked > 0:
        flash(f"Locked {newly_locked} invoice(s). {already_locked} invoice(s) were already locked.", "success")
    elif newly_locked > 0:
        flash(f"Locked {newly_locked} invoice(s).", "success")
    elif already_locked > 0:
        flash(f"{already_locked} invoice(s) are already locked. No changes were made.", "info")
    elif skipped_missing_no > 0:
        flash(f"No invoices could be locked. {skipped_missing_no} invoice(s) are missing invoice numbers.", "warning")
    else:
        flash("No invoices could be locked.", "warning")
    
    return redirect(url_for("invoices.manage_invoices"))


@invoices_bp.route("/invoices/batch-unlock", methods=["POST"])
@login_required
def batch_unlock():
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = [int(x) for x in ids_raw.split(",") if x.strip()] if ids_raw else []
    
    if not invoice_ids:
        flash("No invoices selected", "warning")
        return redirect(url_for("invoices.manage_invoices"))
    
    newly_unlocked = 0
    already_unlocked = 0
    
    for invoice_id in invoice_ids:
        invoice = db.session.get(Invoice, invoice_id)
        if invoice:
            if invoice.locked:
                invoice.locked = False
                newly_unlocked += 1
            else:
                already_unlocked += 1
    
    db.session.commit()
    
    if newly_unlocked > 0 and already_unlocked > 0:
        flash(f"Unlocked {newly_unlocked} invoice(s). {already_unlocked} invoice(s) were already unlocked.", "success")
    elif newly_unlocked > 0:
        flash(f"Unlocked {newly_unlocked} invoice(s).", "success")
    else:
        flash(f"{already_unlocked} invoice(s) are already unlocked. No changes were made.", "info")
    
    return redirect(url_for("invoices.manage_invoices"))


@invoices_bp.route("/invoices/batch-delete", methods=["POST"])
@login_required
def batch_delete():
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = [int(x) for x in ids_raw.split(",") if x.strip()] if ids_raw else []
    
    if not invoice_ids:
        flash("No invoices selected", "warning")
        return redirect(url_for("invoices.manage_invoices"))
    
    deleted_count = 0
    locked_skipped = 0
    has_credit_notes_skipped = 0
    
    for invoice_id in invoice_ids:
        invoice = db.session.get(Invoice, invoice_id)
        if invoice:
            if invoice.locked:
                locked_skipped += 1
            elif CreditNote.query.filter_by(invoice_id=invoice_id).count() > 0:
                has_credit_notes_skipped += 1
            else:
                db.session.delete(invoice)
                deleted_count += 1
    
    db.session.commit()
    
    msg_parts = []
    if deleted_count > 0:
        msg_parts.append(f"Deleted {deleted_count} invoice(s)")
    if locked_skipped > 0:
        msg_parts.append(f"{locked_skipped} locked")
    if has_credit_notes_skipped > 0:
        msg_parts.append(f"{has_credit_notes_skipped} have credit notes")
    
    flash(".".join(msg_parts) + ".", "success")
    
    return redirect(url_for("invoices.manage_invoices"))


@invoices_bp.route("/batch/export/excel", methods=["POST"])
@login_required
def batch_export_excel():
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = [x.strip() for x in ids_raw.split(",") if x.strip()]

    if not invoice_ids:
        flash("No invoices selected", "warning")
        return redirect(url_for("invoices.manage_invoices"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["Invoice Date", "Invoice Number", "GSTIN", "Party Name", "Taxable Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    exported_count = 0
    for invoice_id in invoice_ids:
        invoice = db.session.get(Invoice, int(invoice_id))
        if not invoice or not invoice.invoice_no or not invoice.locked:
            continue

        gst_data = invoice.calculate_gst()
        taxable = gst_data.get("subtotal", 0) or 0

        ws.cell(row=row_idx, column=1, value=invoice.invoice_date.strftime("%d-%m-%Y") if invoice.invoice_date else "")
        ws.cell(row=row_idx, column=2, value=invoice.invoice_no)
        ws.cell(row=row_idx, column=3, value=invoice.party.gstin if invoice.party else "")
        ws.cell(row=row_idx, column=4, value=invoice.party.name if invoice.party else "")
        ws.cell(row=row_idx, column=5, value=taxable)
        row_idx += 1
        exported_count += 1

    if exported_count == 0:
        flash("No locked invoices with invoice numbers to export", "warning")
        return redirect(url_for("invoices.manage_invoices"))

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"invoices_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return make_response(
        send_file(
            output,
            as_attachment=False,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )


@invoices_bp.route("/import/invoices", methods=["GET", "POST"])
@login_required
def import_invoices():
    if request.method == "GET":
        return render_template("import_invoices.html")

    if request.method == "POST":
        file = request.files.get("csv_file")
        if not file or file.filename == "":
            flash("No file selected", "danger")
            return redirect(url_for("invoices.import_invoices"))

        if not file.filename.endswith(".csv"):
            flash("Please upload a CSV file", "danger")
            return redirect(url_for("invoices.import_invoices"))

        import io, csv
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)

        created = 0
        updated = 0
        errors = []
        
        def clean_numeric(value_str):
            if not value_str:
                return None
            cleaned = ''.join(c for c in value_str if c.isdigit() or c in '.-')
            return cleaned if cleaned else None

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                party_gstin = (row.get("party_gstin") or "").strip().upper()
                invoice_no = (row.get("invoice_no") or "").strip()
                invoice_date_str = (row.get("invoice_date") or "").strip()
                description = (row.get("description") or "").strip()
                taxable_value_str = (row.get("taxable_value") or "0").strip()
                igst_rate_str = (row.get("igst_rate") or "0").strip()
                cgst_rate_str = (row.get("cgst_rate") or "0").strip()
                sgst_rate_str = (row.get("sgst_rate") or "0").strip()
                reference_serial_no = (row.get("reference_serial_no") or "").strip()
                reverse_charge_str = (row.get("reverse_charge") or "0").strip()
                is_rcm_str = (row.get("is_rcm") or "0").strip()
                distributor_code = (row.get("distributor_code") or "").strip()
                
                if not party_gstin or not invoice_date_str:
                    errors.append(f"Row {row_num}: Missing party GSTIN or invoice date")
                    continue
                
                if not description:
                    errors.append(f"Row {row_num}: Missing description")
                    continue
                
                if not taxable_value_str:
                    errors.append(f"Row {row_num}: Missing taxable_value")
                    continue
                
                party = Party.query.filter_by(gstin=party_gstin).first()
                if not party:
                    errors.append(f"Row {row_num}: Party with GSTIN {party_gstin} not found")
                    continue
                
                party_name = party.name
                party_address = party.address
                party_gstin_val = party.gstin
                party_pan = party.pan
                party_state = party.state
                party_state_code = party.state_code
                
                invoice_date = None
                for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y"):
                    try:
                        invoice_date = datetime.strptime(invoice_date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                if invoice_date is None:
                    errors.append(f"Row {row_num}: Invalid date format '{invoice_date_str}'")
                    continue
                
                taxable_value = Decimal(clean_numeric(taxable_value_str)) if clean_numeric(taxable_value_str) else Decimal("0")
                igst_rate = float(clean_numeric(igst_rate_str)) if clean_numeric(igst_rate_str) else 0.0
                cgst_rate = float(clean_numeric(cgst_rate_str)) if clean_numeric(cgst_rate_str) else 0.0
                sgst_rate = float(clean_numeric(sgst_rate_str)) if clean_numeric(sgst_rate_str) else 0.0
                reverse_charge = Decimal(clean_numeric(reverse_charge_str)) if clean_numeric(reverse_charge_str) else Decimal("0")
                is_rcm = is_rcm_str in ("1", "true", "True", "yes", "Yes")
                
                tax_type = (row.get("tax_type") or "INTER").strip().upper()
                
                if tax_type == "INTER":
                    igst_amt = (taxable_value * Decimal(str(igst_rate)) / 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    cgst_amt = Decimal("0")
                    sgst_amt = Decimal("0")
                else:
                    igst_amt = Decimal("0")
                    cgst_amt = (taxable_value * Decimal(str(cgst_rate)) / 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    sgst_amt = (taxable_value * Decimal(str(sgst_rate)) / 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                
                final_invoice_no = invoice_no if invoice_no and invoice_no.strip() else None
                
                company = get_current_company()
                
                existing_invoice = None
                if final_invoice_no:
                    existing_invoice = Invoice.query.filter_by(invoice_no=final_invoice_no).first()
                
                if existing_invoice:
                    if existing_invoice.locked:
                        errors.append(f"Row {row_num}: Invoice {final_invoice_no} is locked, skipping update")
                        continue
                    
                    existing_invoice.invoice_date = invoice_date
                    existing_invoice.party_id = party.id
                    existing_invoice.reference_serial_no = reference_serial_no if reference_serial_no else None
                    existing_invoice.tax_type = tax_type
                    existing_invoice.place_of_supply = row.get("place_of_supply")
                    existing_invoice.sac_hsn_code = row.get("sac_hsn_code")
                    existing_invoice.reverse_charge = reverse_charge
                    existing_invoice.is_rcm = is_rcm
                    existing_invoice.distributor_code = distributor_code if distributor_code else None
                    existing_invoice.party_name = party_name
                    existing_invoice.party_address = party_address
                    existing_invoice.party_gstin = party_gstin_val
                    existing_invoice.party_pan = party_pan
                    existing_invoice.party_state = party_state
                    existing_invoice.party_state_code = party_state_code
                    existing_invoice.company_name = company.name if company else "Not Set"
                    existing_invoice.company_address = company.address if company else ""
                    existing_invoice.company_gstin = company.gstin if company else ""
                    existing_invoice.company_pan = company.pan if company else ""
                    
                    InvoiceItem.query.filter_by(invoice_id=existing_invoice.id).delete()
                    invoice = existing_invoice
                else:
                    invoice = Invoice(
                        invoice_no=final_invoice_no,
                        invoice_date=invoice_date,
                        party_id=party.id,
                        reference_serial_no=reference_serial_no if reference_serial_no else None,
                        tax_type=tax_type,
                        place_of_supply=row.get("place_of_supply"),
                        sac_hsn_code=row.get("sac_hsn_code"),
                        reverse_charge=reverse_charge,
                        is_rcm=is_rcm,
                        distributor_code=distributor_code if distributor_code else None,
                        locked=False,
                        party_name=party_name,
                        party_address=party_address,
                        party_gstin=party_gstin_val,
                        party_pan=party_pan,
                        party_state=party_state,
                        party_state_code=party_state_code,
                        company_name=company.name if company else "Not Set",
                        company_address=company.address if company else "",
                        company_gstin=company.gstin if company else "",
                        company_pan=company.pan if company else ""
                    )
                    db.session.add(invoice)
                    db.session.flush()
                
                item = InvoiceItem(
                    invoice_id=invoice.id,
                    description=description,
                    sac_hsn_code=row.get("sac_hsn_code"),
                    taxable_value=taxable_value,
                    igst_rate=igst_rate,
                    igst_amt=igst_amt,
                    cgst_rate=cgst_rate,
                    cgst_amt=cgst_amt,
                    sgst_rate=sgst_rate,
                    sgst_amt=sgst_amt
                )
                db.session.add(item)
                
                invoice.total_in_words = number_to_words(float(taxable_value + igst_amt + cgst_amt + sgst_amt))
                
                if existing_invoice:
                    updated += 1
                else:
                    created += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        if created > 0 or updated > 0:
            db.session.commit()
            if created > 0 and updated > 0:
                flash(f"Created {created} invoice(s), updated {updated} invoice(s)", "success")
            elif created > 0:
                flash(f"Successfully imported {created} invoice(s)", "success")
            else:
                flash(f"Successfully updated {updated} invoice(s)", "success")

        if errors:
            for error in errors[:10]:
                flash(error, "warning")

        return redirect(url_for("invoices.manage_invoices"))

    return render_template("import_invoices.html")

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
import io, csv, os, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pdfkit
from models import CreditNote, CreditNoteItem, Invoice, Party, Company, db
from utils import login_required, get_current_company
from decimal import Decimal, ROUND_HALF_UP

credit_notes_bp = Blueprint('credit_notes', __name__)

@credit_notes_bp.route("/credit-notes", methods=["GET"])
@credit_notes_bp.route("/credit-notes/", methods=["GET"])
@login_required
def manage_credit_notes():
    year = request.args.get("year")
    month = request.args.get("month")
    # Only set defaults if not explicitly provided (for first load)
    if year is None:
        year = str(datetime.now().year)
    if month is None:
        month = str(datetime.now().month).zfill(2)
    search = request.args.get("search", "").strip()
    sort_by = request.args.get("sort_by", "date")
    sort_dir = request.args.get("sort_dir", "desc")
    selected_year = year
    selected_month = month
    search_query = search
    invoice_filter = request.args.get("invoice", "")

    query = CreditNote.query.join(Invoice, CreditNote.invoice_id == Invoice.id).join(Party, Invoice.party_id == Party.id)

    if year:
        query = query.filter(db.extract("year", CreditNote.credit_note_date) == int(year))
    if month:
        query = query.filter(db.extract("month", CreditNote.credit_note_date) == int(month))
    if search:
        query = query.filter(
            db.or_(
                CreditNote.credit_note_no.ilike(f"%{search}%"),
                Invoice.invoice_no.ilike(f"%{search}%"),
                Party.gstin.ilike(f"%{search}%"),
                Party.name.ilike(f"%{search}%"),
            )
        )
    if invoice_filter:
        query = query.filter(CreditNote.invoice_id == int(invoice_filter))

    sort_column = CreditNote.credit_note_date
    if sort_by == "credit_note_no":
        sort_column = CreditNote.credit_note_no
    elif sort_by == "invoice":
        sort_column = Invoice.invoice_no

    if sort_dir == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    credit_notes_list = query.all()
    invoices_list = Invoice.query.filter(Invoice.invoice_no.isnot(None)).order_by(Invoice.invoice_date.desc()).all()
    pending_count = CreditNote.query.filter(CreditNote.credit_note_no.is_(None)).count()

    all_credit_notes = CreditNote.query.all()
    years = sorted(set(cn.credit_note_date.year for cn in all_credit_notes if cn.credit_note_date))
    if not years:
        years = [datetime.now().year]

    return render_template(
        "credit_note_management.html",
        credit_notes=credit_notes_list,
        invoices=invoices_list,
        selected_year=selected_year,
        selected_month=selected_month,
        search_query=search_query,
        years=years,
        sort_by=sort_by,
        sort_dir=sort_dir,
        invoice_filter=invoice_filter,
        pending_count=pending_count,
    )


@credit_notes_bp.route("/credit-note/create", methods=["GET", "POST"])
@login_required
def create_credit_note():
    invoice_id = request.args.get("invoice_id")
    if request.method == "POST":
        invoice_id = request.form.get("invoice_id")
        
        if not invoice_id:
            flash("Invoice selection is required.", "danger")
            return redirect(url_for("credit_notes.create_credit_note"))

        invoice = Invoice.query.get_or_404(invoice_id)
        
        credit_note = CreditNote(
            credit_note_date=datetime.strptime(request.form.get("credit_note_date"), "%Y-%m-%d").date(),
            invoice_id=invoice_id,
            reason=request.form.get("reason"),
            tax_type=request.form.get("tax_type"),
            place_of_supply=request.form.get("place_of_supply"),
        )

        credit_note.party_name = invoice.party_name
        credit_note.party_address = invoice.party_address
        credit_note.party_gstin = invoice.party_gstin
        credit_note.party_pan = invoice.party_pan
        credit_note.party_state = invoice.party_state
        credit_note.party_state_code = invoice.party_state_code

        company_id = request.form.get("company_id")
        if company_id:
            company = db.session.get(Company, company_id)
            if company:
                credit_note.company_name = company.name
                credit_note.company_address = company.address
                credit_note.company_gstin = company.gstin
                credit_note.company_pan = company.pan

        db.session.add(credit_note)
        db.session.flush()

        items_desc = request.form.getlist("item_description[]")
        items_taxable = request.form.getlist("item_taxable_value[]")
        items_cgst_rate = request.form.getlist("item_cgst_rate[]")
        items_sgst_rate = request.form.getlist("item_sgst_rate[]")
        items_igst_rate = request.form.getlist("item_igst_rate[]")

        for i in range(len(items_desc)):
            if items_desc[i].strip():
                taxable = float(items_taxable[i]) if i < len(items_taxable) else 0
                cgst_rate = float(items_cgst_rate[i]) if i < len(items_cgst_rate) else 0
                sgst_rate = float(items_sgst_rate[i]) if i < len(items_sgst_rate) else 0
                igst_rate = float(items_igst_rate[i]) if i < len(items_igst_rate) else 0

                taxable_dec = Decimal(str(taxable))
                if credit_note.tax_type == "INTRA":
                    cgst_amt = float((taxable_dec * Decimal(str(cgst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                    sgst_amt = float((taxable_dec * Decimal(str(sgst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                    igst_amt = 0.0
                elif credit_note.tax_type == "INTER":
                    cgst_amt = 0.0
                    sgst_amt = 0.0
                    igst_amt = float((taxable_dec * Decimal(str(igst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                else:
                    cgst_amt = sgst_amt = igst_amt = 0.0

                item = CreditNoteItem(
                    credit_note_id=credit_note.id,
                    description=items_desc[i],
                    taxable_value=taxable,
                    cgst_rate=cgst_rate,
                    cgst_amt=cgst_amt,
                    sgst_rate=sgst_rate,
                    sgst_amt=sgst_amt,
                    igst_rate=igst_rate,
                    igst_amt=igst_amt,
                )
                db.session.add(item)

        db.session.commit()
        flash("Credit Note created successfully", "success")
        return redirect(url_for("credit_notes.manage_credit_notes"))

    invoices = Invoice.query.filter(Invoice.locked == True).order_by(Invoice.invoice_date.desc()).all()
    companies = Company.query.order_by(Company.name).all()
    
    selected_invoice = None
    invoice_items = None
    selected_company = None
    if invoice_id:
        selected_invoice = db.session.get(Invoice, invoice_id)
        if selected_invoice:
            invoice_items = selected_invoice.items
            if selected_invoice.company_name:
                selected_company = Company.query.filter_by(name=selected_invoice.company_name).first()
    
    return render_template(
        "create_credit_note.html",
        invoices=invoices,
        companies=companies,
        invoice_id=invoice_id,
        selected_invoice=selected_invoice,
        selected_company=selected_company,
        invoice_items=invoice_items
    )


@credit_notes_bp.route("/credit-note/edit/<int:credit_note_id>", methods=["GET", "POST"])
@login_required
def edit_credit_note(credit_note_id):
    credit_note = CreditNote.query.get_or_404(credit_note_id)
    
    if credit_note.locked:
        flash("Locked credit notes cannot be edited.", "warning")
        return redirect(url_for("credit_notes.manage_credit_notes"))
    
    if request.method == "POST":
        invoice_id = request.form.get("invoice_id")
        
        if not invoice_id:
            flash("Invoice selection is required.", "danger")
            return redirect(url_for("credit_notes.edit_credit_note", credit_note_id=credit_note_id))
        
        invoice = Invoice.query.get_or_404(invoice_id)
        
        credit_note.credit_note_date = datetime.strptime(request.form.get("credit_note_date"), "%Y-%m-%d").date()
        credit_note.invoice_id = invoice_id
        credit_note.reason = request.form.get("reason")
        credit_note.tax_type = request.form.get("tax_type")
        credit_note.place_of_supply = request.form.get("place_of_supply")
        
        credit_note.party_name = invoice.party_name
        credit_note.party_address = invoice.party_address
        credit_note.party_gstin = invoice.party_gstin
        credit_note.party_pan = invoice.party_pan
        credit_note.party_state = invoice.party_state
        credit_note.party_state_code = invoice.party_state_code
        
        company_id = request.form.get("company_id")
        if company_id:
            company = db.session.get(Company, company_id)
            if company:
                credit_note.company_name = company.name
                credit_note.company_address = company.address
                credit_note.company_gstin = company.gstin
                credit_note.company_pan = company.pan
        
        for item in credit_note.items:
            db.session.delete(item)
        
        items_desc = request.form.getlist("item_description[]")
        items_taxable = request.form.getlist("item_taxable_value[]")
        items_cgst_rate = request.form.getlist("item_cgst_rate[]")
        items_sgst_rate = request.form.getlist("item_sgst_rate[]")
        items_igst_rate = request.form.getlist("item_igst_rate[]")
        
        for i in range(len(items_desc)):
            if items_desc[i].strip():
                taxable = float(items_taxable[i]) if i < len(items_taxable) else 0
                cgst_rate = float(items_cgst_rate[i]) if i < len(items_cgst_rate) else 0
                sgst_rate = float(items_sgst_rate[i]) if i < len(items_sgst_rate) else 0
                igst_rate = float(items_igst_rate[i]) if i < len(items_igst_rate) else 0
                
                taxable_dec = Decimal(str(taxable))
                if credit_note.tax_type == "INTRA":
                    cgst_amt = float((taxable_dec * Decimal(str(cgst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                    sgst_amt = float((taxable_dec * Decimal(str(sgst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                    igst_amt = 0.0
                elif credit_note.tax_type == "INTER":
                    cgst_amt = 0.0
                    sgst_amt = 0.0
                    igst_amt = float((taxable_dec * Decimal(str(igst_rate)) / Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                else:
                    cgst_amt = sgst_amt = igst_amt = 0.0
                
                item = CreditNoteItem(
                    credit_note_id=credit_note.id,
                    description=items_desc[i],
                    taxable_value=taxable,
                    cgst_rate=cgst_rate,
                    cgst_amt=cgst_amt,
                    sgst_rate=sgst_rate,
                    sgst_amt=sgst_amt,
                    igst_rate=igst_rate,
                    igst_amt=igst_amt,
                )
                db.session.add(item)
        
        db.session.commit()
        flash("Credit Note updated successfully", "success")
        return redirect(url_for("credit_notes.manage_credit_notes"))
    
    invoices = Invoice.query.filter(Invoice.locked == True).order_by(Invoice.invoice_date.desc()).all()
    companies = Company.query.order_by(Company.name).all()
    return render_template("edit_credit_note.html", credit_note=credit_note, invoices=invoices, companies=companies)


@credit_notes_bp.route("/credit-note/preview/<int:credit_note_id>")
@login_required
def preview_credit_note(credit_note_id):
    credit_note = CreditNote.query.get_or_404(credit_note_id)
    return render_template("credit_note_preview.html", credit_note=credit_note)


@credit_notes_bp.route("/credit-note/delete/<int:credit_note_id>")
@login_required
def delete_credit_note(credit_note_id):
    credit_note = CreditNote.query.get_or_404(credit_note_id)
    db.session.delete(credit_note)
    db.session.commit()
    flash("Credit Note deleted successfully", "success")
    return redirect(url_for("credit_notes.manage_credit_notes"))


@credit_notes_bp.route("/credit-note/pdf/<int:credit_note_id>")
@login_required
def generate_credit_note_pdf(credit_note_id):
    from flask import current_app
    credit_note = CreditNote.query.get_or_404(credit_note_id)
    
    def sanitize(name):
        name = str(name).strip()
        name = re.sub(r'[<>:"/\\|?*&]', "_", name)
        return re.sub(r"_+", "_", name).strip("_")
    
    settings = {
        "company_name": credit_note.company_name or "Not set",
        "address": credit_note.company_address or "",
        "gstin": credit_note.company_gstin or "",
        "pan": credit_note.company_pan or "",
        "logo": "",
        "place_of_supply": credit_note.place_of_supply or "",
        "state_code": "",
    }
    html_content = render_template("credit_note_pdf_template.html", credit_note=credit_note, settings=settings)
    
    export_folder = current_app.config["EXPORT_FOLDER"]
    os.makedirs(export_folder, exist_ok=True)
    
    cn_no = sanitize(credit_note.credit_note_no or f"cn_{credit_note.id}")
    pdf_path = os.path.join(export_folder, f"credit_note_{cn_no}.pdf")
    pdfkit.from_string(html_content, pdf_path)
    
    return send_file(pdf_path, as_attachment=False, mimetype='application/pdf')


@credit_notes_bp.route("/credit-note/generate-numbers", methods=["POST"])
@login_required
def generate_credit_note_numbers_route():
    current_fy = f"{datetime.now().year}-{datetime.now().year + 1}"
    fy_prefix = current_fy.split("-")[0][-2:] + "-" + current_fy[-2:]

    pending = CreditNote.query.filter(CreditNote.credit_note_no.is_(None)).order_by(CreditNote.credit_note_date).all()

    if not pending:
        flash("No pending credit notes to generate numbers for", "warning")
        return redirect(url_for("credit_notes.manage_credit_notes"))

    last_cn = (
        CreditNote.query
        .filter(CreditNote.credit_note_no.isnot(None))
        .filter(CreditNote.credit_note_no.like(f"CRN/{fy_prefix}%"))
        .order_by(CreditNote.credit_note_no.desc())
        .first()
    )

    seq = int(last_cn.credit_note_no.split("/")[2]) if last_cn else 0

    for cn in pending:
        seq += 1
        cn.credit_note_no = f"CRN/{fy_prefix}/{str(seq).zfill(3)}"

    db.session.commit()
    flash(f"Generated numbers for {len(pending)} credit notes", "success")
    return redirect(url_for("credit_notes.manage_credit_notes"))


@credit_notes_bp.route("/credit-notes/batch-lock", methods=["POST"])
@login_required
def batch_lock_credit_notes():
    ids_raw = request.form.get("credit_note_ids", "")
    if not ids_raw:
        flash("No credit notes selected", "warning")
        return redirect(url_for("credit_notes.manage_credit_notes"))
    
    ids = [int(x) for x in ids_raw.split(",") if x.strip()]
    CreditNote.query.filter(CreditNote.id.in_(ids)).update({CreditNote.locked: True}, synchronize_session=False)
    db.session.commit()
    flash(f"Locked {len(ids)} credit notes", "success")
    return redirect(url_for("credit_notes.manage_credit_notes"))


@credit_notes_bp.route("/credit-notes/batch-unlock", methods=["POST"])
@login_required
def batch_unlock_credit_notes():
    ids_raw = request.form.get("credit_note_ids", "")
    if not ids_raw:
        flash("No credit notes selected", "warning")
        return redirect(url_for("credit_notes.manage_credit_notes"))
    
    ids = [int(x) for x in ids_raw.split(",") if x.strip()]
    CreditNote.query.filter(CreditNote.id.in_(ids)).update({CreditNote.locked: False}, synchronize_session=False)
    db.session.commit()
    flash(f"Unlocked {len(ids)} credit notes", "success")
    return redirect(url_for("credit_notes.manage_credit_notes"))


@credit_notes_bp.route("/credit-notes/batch-delete", methods=["POST"])
@login_required
def batch_delete_credit_notes():
    ids_raw = request.form.get("credit_note_ids", "")
    if not ids_raw:
        flash("No credit notes selected", "warning")
        return redirect(url_for("credit_notes.manage_credit_notes"))
    
    ids = [int(x) for x in ids_raw.split(",") if x.strip()]
    credit_notes = CreditNote.query.filter(CreditNote.id.in_(ids)).all()
    deleted_count = 0
    for cn in credit_notes:
        db.session.delete(cn)
        deleted_count += 1
    db.session.commit()
    flash(f"Deleted {deleted_count} credit notes", "success")
    return redirect(url_for("credit_notes.manage_credit_notes"))


@credit_notes_bp.route("/credit-notes/batch-export", methods=["POST"])
@login_required
def batch_export_credit_notes():
    import zipfile
    ids_raw = request.form.get("credit_note_ids", "")
    if not ids_raw:
        flash("No credit notes selected", "warning")
        return redirect(url_for("credit_notes.manage_credit_notes"))
    
    ids = [int(x) for x in ids_raw.split(",") if x.strip()]
    
    credit_notes = CreditNote.query.filter(CreditNote.id.in_(ids)).all()
    valid_cns = [cn for cn in credit_notes if cn.locked and cn.credit_note_no]
    
    if not valid_cns:
        flash("No locked credit notes with credit note numbers to export", "warning")
        return redirect(url_for("credit_notes.manage_credit_notes"))
    
    from utils import get_current_company
    company = get_current_company()
    settings = {
        "company_name": company.name if company else "Not Set",
        "address": company.address if company else "",
        "gstin": company.gstin if company else "",
        "pan": company.pan if company else "",
    }
    
    os.makedirs("temp_exports", exist_ok=True)
    zip_path = "temp_exports/credit_notes_batch.zip"
    
    def sanitize(name):
        name = str(name).strip()
        name = re.sub(r'[<>:"/\\|?*&]', "_", name)
        return re.sub(r"_+", "_", name).strip("_")
    
    with zipfile.ZipFile(zip_path, "w") as zipf:
        for cn in valid_cns:
            html = render_template("credit_note_pdf_template.html", credit_note=cn, settings=settings)
            pdf = pdfkit.from_string(html, False, options={"enable-local-file-access": ""})
            
            safe_name = f"CN_{sanitize(cn.credit_note_no)}.pdf"
            zipf.writestr(safe_name, pdf)
    
    return send_file(zip_path, as_attachment=False, download_name=f"credit_notes_batch.zip")


@credit_notes_bp.route("/credit-notes/batch-export-excel", methods=["POST"])
@login_required
def batch_export_excel_credit_notes():
    ids_raw = request.form.get("credit_note_ids", "")
    if not ids_raw:
        flash("No credit notes selected", "warning")
        return redirect(url_for("credit_notes.manage_credit_notes"))
    
    ids = [int(x) for x in ids_raw.split(",") if x.strip()]
    
    credit_notes = CreditNote.query.filter(CreditNote.id.in_(ids)).all()
    valid_cns = [cn for cn in credit_notes if cn.locked and cn.credit_note_no]
    
    if not valid_cns:
        flash("No locked credit notes with credit note numbers to export", "warning")
        return redirect(url_for("credit_notes.manage_credit_notes"))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Credit Notes"
    
    headers = ["CN Date", "CN No.", "Party GSTIN", "Party Name", "Taxable", "IGST", "CGST", "SGST", "Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    for row_idx, cn in enumerate(valid_cns, 2):
        gst = cn.calculate_gst()
        ws.cell(row=row_idx, column=1, value=cn.credit_note_date.strftime("%d-%m-%Y") if cn.credit_note_date else "")
        ws.cell(row=row_idx, column=2, value=cn.credit_note_no)
        ws.cell(row=row_idx, column=3, value=cn.party_gstin or "")
        ws.cell(row=row_idx, column=4, value=cn.party_name or "")
        ws.cell(row=row_idx, column=5, value=gst.get("subtotal", 0) or 0)
        ws.cell(row=row_idx, column=6, value=gst.get("igst", 0) or 0)
        ws.cell(row=row_idx, column=7, value=gst.get("cgst", 0) or 0)
        ws.cell(row=row_idx, column=8, value=gst.get("sgst", 0) or 0)
        ws.cell(row=row_idx, column=9, value=gst.get("total", 0) or 0)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name="credit_notes.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@credit_notes_bp.route("/credit-notes/import", methods=["GET", "POST"])
@login_required
def import_credit_notes():
    if request.method == "GET":
        return render_template("import_credit_notes.html")
    
    file = request.files.get("csv_file")
    if not file or file.filename == "":
        flash("No file selected", "danger")
        return redirect(url_for("credit_notes.import_credit_notes"))
    
    if not file.filename.endswith(".csv"):
        flash("Please upload a CSV file", "danger")
        return redirect(url_for("credit_notes.import_credit_notes"))
    
    stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
    csv_reader = csv.DictReader(stream)
    
    imported = 0
    errors = []
    
    def clean_numeric(value_str):
        if not value_str:
            return None
        cleaned = ''.join(c for c in value_str if c.isdigit() or c in '.-')
        return cleaned if cleaned else None
    
    for row_num, row in enumerate(csv_reader, start=2):
        try:
            party_gstin = (row.get("party_gstin") or "").strip().upper()
            credit_note_date_str = (row.get("credit_note_date") or "").strip()
            reference_invoice_no = (row.get("reference_invoice_number") or "").strip()
            description = (row.get("description") or "").strip()
            taxable_value_str = (row.get("taxable_value") or "0").strip()
            igst_rate_str = (row.get("igst_rate") or "0").strip()
            cgst_rate_str = (row.get("cgst_rate") or "0").strip()
            sgst_rate_str = (row.get("sgst_rate") or "0").strip()
            reason = (row.get("reason") or "").strip()
            place_of_supply = (row.get("place_of_supply") or "").strip()
            
            if not party_gstin or not credit_note_date_str:
                errors.append(f"Row {row_num}: Missing party GSTIN or credit note date")
                continue
            
            if not description:
                errors.append(f"Row {row_num}: Missing description")
                continue
            
            if not taxable_value_str:
                errors.append(f"Row {row_num}: Missing taxable_value")
                continue
            
            if not reference_invoice_no:
                errors.append(f"Row {row_num}: Missing reference invoice number")
                continue
            
            invoice = Invoice.query.filter_by(invoice_no=reference_invoice_no).first()
            if not invoice:
                errors.append(f"Row {row_num}: Invoice with number {reference_invoice_no} not found")
                continue
            
            party = Party.query.filter_by(gstin=party_gstin).first()
            if not party:
                errors.append(f"Row {row_num}: Party with GSTIN {party_gstin} not found")
                continue
            
            credit_note_date = None
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y"):
                try:
                    credit_note_date = datetime.strptime(credit_note_date_str, fmt).date()
                    break
                except ValueError:
                    continue
            if credit_note_date is None:
                errors.append(f"Row {row_num}: Invalid date format '{credit_note_date_str}'")
                continue
            
            taxable_value = Decimal(clean_numeric(taxable_value_str)) if clean_numeric(taxable_value_str) else Decimal("0")
            igst_rate = float(clean_numeric(igst_rate_str)) if clean_numeric(igst_rate_str) else 0.0
            cgst_rate = float(clean_numeric(cgst_rate_str)) if clean_numeric(cgst_rate_str) else 0.0
            sgst_rate = float(clean_numeric(sgst_rate_str)) if clean_numeric(sgst_rate_str) else 0.0
            
            tax_type = (row.get("tax_type") or "INTER").strip().upper()
            
            if tax_type == "INTER":
                igst_amt = (taxable_value * Decimal(str(igst_rate)) / 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                cgst_amt = Decimal("0")
                sgst_amt = Decimal("0")
            else:
                igst_amt = Decimal("0")
                cgst_amt = (taxable_value * Decimal(str(cgst_rate)) / 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                sgst_amt = (taxable_value * Decimal(str(sgst_rate)) / 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            
            credit_note = CreditNote(
                credit_note_date=credit_note_date,
                invoice_id=invoice.id,
                reason=reason,
                tax_type=tax_type,
                place_of_supply=place_of_supply,
                locked=False,
                party_name=party.name,
                party_address=party.address,
                party_gstin=party.gstin,
                party_pan=party.pan,
                party_state=party.state,
                party_state_code=party.state_code,
                company_name=invoice.company_name,
                company_address=invoice.company_address,
                company_gstin=invoice.company_gstin,
                company_pan=invoice.company_pan,
            )
            
            db.session.add(credit_note)
            db.session.flush()
            
            item = CreditNoteItem(
                credit_note_id=credit_note.id,
                description=description,
                taxable_value=taxable_value,
                igst_rate=igst_rate,
                igst_amt=igst_amt,
                cgst_rate=cgst_rate,
                cgst_amt=cgst_amt,
                sgst_rate=sgst_rate,
                sgst_amt=sgst_amt,
            )
            db.session.add(item)
            
            imported += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")
    
    if imported > 0:
        db.session.commit()
        flash(f"Successfully imported {imported} credit notes", "success")
    
    if errors:
        for error in errors[:10]:
            flash(error, "warning")
    
    return redirect(url_for("credit_notes.manage_credit_notes"))
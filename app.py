import os
import csv
import io
import zipfile
import secrets
from datetime import datetime, timedelta, timezone
from functools import wraps
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    send_file,
    jsonify,
    make_response,
    session,
    g,
)
from werkzeug.utils import secure_filename
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from flask_sqlalchemy import SQLAlchemy
from models import (
    db,
    Settings,
    Party,
    Invoice,
    InvoiceItem,
    CreditNote,
    CreditNoteItem,
    User,
    RecoveryCode,
    ConfigStore,
    hash_password,
    generate_recovery_code,
)
from helpers import (
    get_date_range,
    get_last_month_dates,
    get_two_months_back_dates,
    get_three_months_back_dates,
    get_four_months_back_dates,
    get_last_3_months_dates,
    get_this_month_last_year_dates,
    get_filtered_invoices,
    calculate_revenue_and_gst,
    get_prefiltered_invoices,
    precalculate_gst_data,
    calculate_party_growth_data,
    sort_party_growth_data,
    add_missing_parties,
    calculate_revenue_change,
    get_pending_invoices_info,
    prepare_chart_data,
    calculate_top_parties,
    get_trend_months,
    get_month_name,
)
from forms import (
    LoginForm,
    SetupForm,
    PartyForm,
    InvoiceForm,
    CreditNoteForm,
    RecoveryCodeForm,
    UserForm,
)

# Constants
RECOVERY_CODE_COUNT = 5
INVOICE_NUMBER_PADDING = 3
MAX_RETRY_ATTEMPTS = 3

app = Flask(__name__)

app.config["SECRET_KEY"] = os.environ.get(
    "SECRET_KEY", "dev-secret-key-change-in-production"
)
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL",
    f"postgresql+psycopg://{os.environ.get('DB_USER', 'postgres')}:{os.environ.get('DB_PASSWORD', 'postgres')}@{os.environ.get('DB_HOST', 'localhost')}:{os.environ.get('DB_PORT', '5432')}/{os.environ.get('DB_NAME', 'gst_invoices')}",
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

app.config["UPLOAD_FOLDER"] = "static/uploads"
app.config["EXPORT_FOLDER"] = "exports"
app.config["LOGO_FOLDER"] = "static/logos"

os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["EXPORT_FOLDER"], exist_ok=True)
os.makedirs(app.config["LOGO_FOLDER"], exist_ok=True)


@app.before_request
def check_auth():
    public_routes = ["login", "setup", "recovery", "static", "index"]
    if request.endpoint and request.endpoint.split(".")[0] in public_routes:
        return
    if not session.get("user_id"):
        return redirect(url_for("login"))


db.init_app(app)


def parse_date(date_str):
    """Parse date from various formats"""
    if not date_str:
        return None
    date_str = date_str.strip()
    formats = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"]
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(value):
    """Parse numeric value from various formats (handles commas, currency symbols)"""
    if not value:
        return 0.0
    value = str(value).strip()
    value = value.replace("?", "").replace("₹", "").replace(",", "").replace("'", "")
    try:
        return float(value)
    except ValueError:
        return 0.0


def parse_percentage(value):
    """Parse percentage value (handles % sign)"""
    if not value:
        return 0.0
    value = str(value).strip().replace("%", "")
    try:
        return float(value)
    except ValueError:
        return 0.0


def parse_tax_type(tax_type):
    """Normalize tax type"""
    if not tax_type:
        return "INTER"
    tax_type = tax_type.strip().upper()
    if tax_type in ["INTRA", "INNERS", "INRA"]:
        return "INTRA"
    return "INTER"


def validate_tax_rates(tax_type, is_rcm, cgst_rate, sgst_rate, igst_rate):
    """
    Validate tax rates based on rules:
    1. RCM only: no igst, cgst, sgst
    2. INTER only: IGST, no cgst, no sgst
    3. INTRA only: CGST + SGST, no igst

    Returns (is_valid, error_message)
    """
    if is_rcm:
        if cgst_rate > 0 or sgst_rate > 0 or igst_rate > 0:
            return False, "RCM enabled - tax rates must be 0"
        return True, None

    if tax_type == "INTER":
        if cgst_rate > 0 or sgst_rate > 0:
            return False, "INTER state - IGST required, CGST/SGST must be 0"
        if igst_rate <= 0:
            return False, "INTER state - IGST rate is required"
        return True, None

    if tax_type == "INTRA":
        if igst_rate > 0:
            return False, "INTRA state - CGST/SGST required, IGST must be 0"
        if cgst_rate <= 0 or sgst_rate <= 0:
            return False, "INTRA state - CGST and SGST rates are required"
        return True, None

    return True, None


def extract_pan_from_gstin(gstin):
    """Extract PAN from GSTIN (digits 3-12)"""
    if not gstin or len(gstin) < 12:
        return ""
    return gstin[2:12].upper()


def number_to_words(num):
    """Convert number to words format: 'Nineteen Thousand Six Hundred Forty Only and Eighty Four Paisa'"""
    num = float(num)
    if num == 0:
        return "Zero Only"

    # Define word arrays
    ones = [
        "",
        "One",
        "Two",
        "Three",
        "Four",
        "Five",
        "Six",
        "Seven",
        "Eight",
        "Nine",
        "Ten",
        "Eleven",
        "Twelve",
        "Thirteen",
        "Fourteen",
        "Fifteen",
        "Sixteen",
        "Seventeen",
        "Eighteen",
        "Nineteen",
    ]
    tens = [
        "",
        "",
        "Twenty",
        "Thirty",
        "Forty",
        "Fifty",
        "Sixty",
        "Seventy",
        "Eighty",
        "Ninety",
    ]
    thousands = ["", "Thousand", "Million", "Billion"]

    def convert_hundreds(n):
        """Convert number less than 1000 to words"""
        words = []
        if n >= 100:
            words.append(ones[n // 100])
            words.append("Hundred")
            n %= 100
        if n >= 20:
            words.append(tens[n // 10])
            n %= 10
        if n > 0:
            words.append(ones[n])
        return " ".join(words)

    def convert_number(n):
        """Convert any number to words"""
        if n == 0:
            return ""

        words = []
        thousand_index = 0

        while n > 0:
            remainder = n % 1000
            if remainder > 0:
                remainder_words = convert_hundreds(remainder)
                if thousand_index > 0:
                    remainder_words += " " + thousands[thousand_index]
                words.insert(0, remainder_words)
            n //= 1000
            thousand_index += 1

        return " ".join(words)

    # Split into rupees and paise
    integer_part = int(num)
    paise = int(round((num - integer_part) * 100))

    # Convert rupees to words
    rupees_words = convert_number(integer_part)
    if rupees_words:
        rupees_words += " Only"
    else:
        rupees_words = "Zero Only"

    # Convert paise to words if needed
    if paise > 0:
        paise_words = convert_number(paise)
        result = f"{rupees_words} and {paise_words} Paisa"
    else:
        result = rupees_words

    return result


def get_fiscal_year():
    now = datetime.now()
    if now.month >= 4:
        return f"{now.year}-{now.year + 1}"
    else:
        return f"{now.year - 1}-{now.year}"


def get_fy_dates(fy):
    parts = fy.split("-")
    if len(parts) != 2:
        return datetime(datetime.now().year, 4, 1), datetime(
            datetime.now().year + 1, 3, 31
        )
    start_year = int(parts[0])
    start_date = datetime(start_year, 4, 1)
    end_date = datetime(start_year + 1, 3, 31)
    return start_date, end_date


def current_month():
    return datetime.now().month


def get_fy_short():
    fy = get_fiscal_year()
    return fy.replace("-", "_")


def get_month_name(month_num):
    months = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]
    return months[month_num - 1] if 1 <= month_num <= 12 else ""


def sanitize_filename(name):
    import re

    name = str(name).strip()
    name = re.sub(r'[<>:"/\\|?*&]', "_", name)
    name = re.sub(r"_+", "_", name)
    return name.strip("_")


def get_export_path(invoice):
    fy_short = get_fy_short()
    month_name = (
        get_month_name(invoice.invoice_date.month)
        if invoice.invoice_date
        else "Unknown"
    )
    party_name = sanitize_filename(
        invoice.party_name or invoice.party.name if invoice.party else "Unknown"
    )
    invoice_no = sanitize_filename(invoice.invoice_no or f"invoice_{invoice.id}")
    return os.path.join(
        app.config["EXPORT_FOLDER"],
        fy_short,
        month_name,
        f"{party_name}_{invoice_no}.pdf",
    )


def get_greeting():
    hour = datetime.now().hour
    if hour < 12:
        return "Good Morning"
    elif hour < 17:
        return "Good Afternoon"
    else:
        return "Good Evening"


def generate_invoice_numbers():
    current_fy = get_fiscal_year()
    fy_prefix = current_fy.split("-")[0][-2:] + "-" + current_fy[-2:]

    pending_invoices = (
        Invoice.query.filter(Invoice.invoice_no.is_(None))
        .order_by(Invoice.invoice_date)
        .all()
    )

    if not pending_invoices:
        return 0

    last_invoice = (
        Invoice.query.filter(Invoice.invoice_no.isnot(None))
        .filter(Invoice.invoice_no.like(f"{fy_prefix}%"))
        .order_by(Invoice.invoice_no.desc())
        .first()
    )

    if last_invoice:
        last_no = last_invoice.invoice_no
        parts = last_no.split("/")
        if len(parts) == 2:
            try:
                seq = int(parts[1])
            except:
                seq = 0
        else:
            seq = 0
    else:
        seq = 0

    for invoice in pending_invoices:
        seq += 1
        invoice_no = f"{fy_prefix}/{str(seq).zfill(3)}"
        invoice.invoice_no = invoice_no

        total = invoice.calculate_gst()["total"]
        invoice.total_in_words = number_to_words(total)

    db.session.commit()
    return len(pending_invoices)


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.url))
        if session.get("role") != "admin":
            flash("Admin access required", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))

    # If no users exist, show setup link
    if not User.query.first():
        session["no_users"] = True

    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.is_active and user.check_password(form.password.data):
            session["user_id"] = user.id
            session["username"] = user.username
            session["full_name"] = user.full_name
            session["role"] = user.role

            session.pop("setup_codes", None)
            session.pop("new_codes", None)

            if form.remember.data:
                session.permanent = True
                app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)
            else:
                session.permanent = True
                app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=8)

            flash(f"Welcome back, {user.full_name}!", "success")
            next_url = request.args.get("next")
            return redirect(next_url or url_for("dashboard"))
        else:
            flash("Invalid username or password", "danger")

    return render_template("login.html", form=form)


@app.route("/logout")
def logout():
    session.pop("setup_codes", None)
    session.pop("new_codes", None)
    session.clear()
    return redirect(url_for("login"))


@app.route("/setup", methods=["GET", "POST"])
def setup():
    if User.query.first():
        return redirect(url_for("login"))

    form = SetupForm()
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            full_name=form.full_name.data,
            role="admin",
            must_change_password=False,
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.flush()

        codes = []
        for _ in range(RECOVERY_CODE_COUNT):
            code = generate_recovery_code()
            codes.append(code)
            rc = RecoveryCode(code=code, user_id=user.id)
            db.session.add(rc)

        db.session.commit()

        session["setup_codes"] = codes
        ConfigStore.set("setup_codes", ",".join(codes))

        flash("Setup complete!", "success")
        return redirect(url_for("setup_success"))

    return render_template("setup.html", form=form)


@app.route("/setup-success")
def setup_success():
    codes = session.get("setup_codes", [])
    if not codes:
        codes_str = ConfigStore.get("setup_codes", "")
        codes = [c for c in codes_str.split(",") if c] if codes_str else []
    if not codes:
        return redirect(url_for("setup"))
    return render_template("setup_success.html", codes=codes)


@app.route("/download-setup-codes")
def download_setup_codes():
    codes = session.get("setup_codes", [])
    if not codes:
        codes_str = ConfigStore.get("setup_codes", "")
        codes = [c for c in codes_str.split(",") if c] if codes_str else []
    if not codes:
        flash("No codes to download", "warning")
        return redirect(url_for("setup"))

    content = "Recovery Codes\n"
    content += "=" * 50 + "\n\n"
    for i, code in enumerate(codes, 1):
        content += f"{i}. {code[:4]}-{code[4:]}\n"
    content += "\n" + "=" * 50 + "\n"
    content += "Each code can only be used once.\n"
    content += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"

    return (
        content.encode(),
        200,
        {
            "Content-Type": "text/plain",
            "Content-Disposition": "attachment;filename=recovery_codes.txt",
        },
    )


@app.route("/recovery", methods=["GET", "POST"])
def recovery():
    if request.method == "POST":
        code = request.form.get("code", "").strip().replace("-", "").replace(" ", "")

        recovery = RecoveryCode.query.filter_by(
            code=code.upper(), is_used=False
        ).first()

        if not recovery:
            flash("Invalid or used recovery code", "danger")
        else:
            user = db.session.get(User, recovery.user_id)
            new_password = request.form.get("password", "")
            confirm = request.form.get("confirm_password", "")

            if not new_password:
                recovery.is_used = True
                recovery.used_at = datetime.now(timezone.utc)
                db.session.commit()
                session["recovery_user_id"] = user.id
                session["resetting_password"] = True
                flash("Code verified! Set your new password.", "success")
                return redirect(url_for("recovery"))

            if new_password != confirm:
                flash("Passwords do not match", "danger")
            elif len(new_password) < 6:
                flash("Password must be at least 6 characters", "danger")
            else:
                user.set_password(new_password)
                recovery.is_used = True
                recovery.used_at = datetime.now(timezone.utc)
                db.session.commit()
                flash("Password reset successful! Please login.", "success")
                return redirect(url_for("login"))

    return render_template("recovery.html")


@app.route("/generate-recovery-codes", methods=["GET", "POST"])
@admin_required
def generate_recovery_codes():
    if request.method == "GET":
        flash("Invalid request method", "warning")
        return redirect(url_for("manage_users"))

    user_id = request.form.get("user_id")
    if not user_id:
        flash("User ID is required", "danger")
        return redirect(url_for("manage_users"))

    user = db.session.get(User, int(user_id))
    if not user:
        flash("User not found", "danger")
        return redirect(url_for("manage_users"))

    # Delete existing unused codes using a more robust method
    RecoveryCode.query.filter_by(user_id=user.id, is_used=False).delete(
        synchronize_session=False
    )
    db.session.commit()

    codes = []
    for _ in range(RECOVERY_CODE_COUNT):
        code = generate_recovery_code()
        codes.append(code)
        rc = RecoveryCode(code=code, user_id=user_id)
        db.session.add(rc)

    db.session.commit()
    session["new_codes"] = codes
    ConfigStore.set("temp_recovery_codes", ",".join(codes))
    flash("New recovery codes generated", "success")
    return redirect(url_for("manage_users"))


@app.route("/download-recovery-codes")
@admin_required
def download_recovery_codes():
    codes_data = ConfigStore.get("temp_recovery_codes")
    if not codes_data:
        flash("No codes to download", "warning")
        return redirect(url_for("manage_users"))

    codes = codes_data.split(",")
    content = "Recovery Codes\n"
    content += "=" * 50 + "\n\n"
    for i, code in enumerate(codes, 1):
        content += f"{i}. {code[:4]}-{code[4:]}\n"
    content += "\n" + "=" * 50 + "\n"
    content += "Each code can only be used once.\n"
    content += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"

    ConfigStore.set("temp_recovery_codes", "")

    return (
        content.encode(),
        200,
        {
            "Content-Type": "text/plain",
            "Content-Disposition": "attachment;filename=recovery_codes.txt",
        },
    )


@app.route("/backup")
@admin_required
def backup():
    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        parties = Party.query.all()
        party_data = io.StringIO()
        writer = csv.writer(party_data)
        writer.writerow(["gstin", "name", "pan", "amc_code", "address", "state", "state_code"])
        for p in parties:
            writer.writerow([p.gstin, p.name, p.pan, p.amc_code, p.address, p.state, p.state_code])
        zf.writestr("parties.csv", party_data.getvalue())

        invoices = Invoice.query.all()
        invoice_data = io.StringIO()
        writer = csv.writer(invoice_data)
        writer.writerow(["invoice_no", "reference_serial_no", "invoice_date", "party_gstin", "description", "sac_hsn_code", "taxable_value", "tax_type", "cgst_rate", "sgst_rate", "igst_rate", "place_of_supply", "reverse_charge", "is_rcm", "distributor_code"])
        for inv in invoices:
            party = db.session.get(Party, inv.party_id)
            party_gstin = party.gstin if party else ""
            items = InvoiceItem.query.filter_by(invoice_id=inv.id).all()
            for item in items:
                cgst_rate = item.cgst_rate or 0
                sgst_rate = item.sgst_rate or 0
                igst_rate = item.igst_rate or 0
                row = [
                    inv.invoice_no or "",
                    inv.reference_serial_no or "",
                    inv.invoice_date.isoformat() if inv.invoice_date else "",
                    party_gstin,
                    item.description or "",
                    inv.sac_hsn_code or "",
                    item.taxable_value,
                    inv.tax_type,
                    cgst_rate,
                    sgst_rate,
                    igst_rate,
                    inv.place_of_supply or "",
                    inv.reverse_charge or 0,
                    "1" if inv.is_rcm else "0",
                    inv.distributor_code or "",
                ]
                writer.writerow(row)
        zf.writestr("invoices.csv", invoice_data.getvalue())

        credit_notes = CreditNote.query.all()
        cn_data = io.StringIO()
        writer = csv.writer(cn_data)
        writer.writerow(["credit_note_no", "credit_note_date", "invoice_no", "reason", "tax_type", "place_of_supply", "description", "sac_hsn_code", "taxable_value", "cgst_rate", "sgst_rate", "igst_rate"])
        for cn in credit_notes:
            invoice = db.session.get(Invoice, cn.invoice_id)
            invoice_no = invoice.invoice_no if invoice else ""
            items = CreditNoteItem.query.filter_by(credit_note_id=cn.id).all()
            for item in items:
                cgst_rate = item.cgst_rate or 0
                sgst_rate = item.sgst_rate or 0
                igst_rate = item.igst_rate or 0
                row = [
                    cn.credit_note_no or "",
                    cn.credit_note_date.isoformat() if cn.credit_note_date else "",
                    invoice_no,
                    cn.reason or "",
                    cn.tax_type,
                    cn.place_of_supply or "",
                    item.description or "",
                    "",
                    item.taxable_value,
                    cgst_rate,
                    sgst_rate,
                    igst_rate,
                ]
                writer.writerow(row)
        zf.writestr("credit_notes.csv", cn_data.getvalue())

        users = User.query.all()
        user_data = io.StringIO()
        writer = csv.writer(user_data)
        writer.writerow(["username", "full_name", "role", "is_active"])
        for u in users:
            writer.writerow([u.username, u.full_name, u.role, "1" if u.is_active else "0"])
        zf.writestr("users.csv", user_data.getvalue())

        settings_data = io.StringIO()
        writer = csv.writer(settings_data)
        writer.writerow(["key", "value"])
        for key in ["company_name", "arn_number", "address", "gstin", "pan", "place_of_supply", "state_code", "logo"]:
            writer.writerow([key, Settings.get(key, "")])
        zf.writestr("settings.csv", settings_data.getvalue())

        manifest = f"Backup created: {datetime.now().isoformat()}\n"
        manifest += f"Parties: {len(parties)}\n"
        manifest += f"Invoices: {len(invoices)}\n"
        manifest += f"Credit Notes: {len(credit_notes)}\n"
        manifest += f"Users: {len(users)}\n"
        zf.writestr("manifest.txt", manifest)

    buffer.seek(0)
    filename = f"gst_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"

    return send_file(
        buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/")
def index():
    if not session.get("user_id"):
        return redirect(url_for("login"))
    if not User.query.first():
        return redirect(url_for("setup"))
    return redirect(url_for("dashboard"))


@app.route("/company", methods=["GET", "POST"])
@login_required
def company():
    if request.method == "POST":
        Settings.set("company_name", request.form.get("company_name"))
        Settings.set("arn_number", request.form.get("arn_number"))
        Settings.set("address", request.form.get("address"))
        Settings.set("gstin", request.form.get("gstin"))
        Settings.set("pan", request.form.get("pan"))
        Settings.set("place_of_supply", request.form.get("place_of_supply"))
        Settings.set("state_code", request.form.get("state_code"))

        if "logo" in request.files and request.files["logo"].filename:
            logo = request.files["logo"]
            logo_path = os.path.join(app.config["LOGO_FOLDER"], "logo.png")
            logo.save(logo_path)
            Settings.set("logo", "logo.png")

        flash("Company settings saved successfully", "success")
        return redirect(url_for("company"))

    company_name = Settings.get("company_name", "")
    arn_number = Settings.get("arn_number", "")
    address = Settings.get("address", "")
    gstin = Settings.get("gstin", "")
    pan = Settings.get("pan", "")
    place_of_supply = Settings.get("place_of_supply", "")
    state_code = Settings.get("state_code", "")
    logo = Settings.get("logo", "")

    return render_template(
        "company.html",
        company_name=company_name,
        arn_number=arn_number,
        address=address,
        gstin=gstin,
        pan=pan,
        place_of_supply=place_of_supply,
        state_code=state_code,
        logo=logo,
    )


@app.route("/delete-logo")
@login_required
def delete_logo():
    logo_filename = Settings.get("logo", "")
    if logo_filename:
        logo_path = os.path.join(app.config["LOGO_FOLDER"], logo_filename)
        if os.path.exists(logo_path):
            os.remove(logo_path)
        Settings.set("logo", "")
    flash("Logo deleted successfully", "success")
    return redirect(url_for("company"))


@app.route("/users", methods=["GET", "POST"])
@admin_required
def manage_users():
    display_codes = session.pop("new_codes", None)

    if request.method == "POST":
        action = request.form.get("action")

        if action == "create":
            full_name = request.form.get("full_name", "").strip()
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            role = request.form.get("role", "staff")

            if not full_name or not username or not password:
                flash("Full name, username and password required", "danger")
            elif User.query.filter_by(username=username).first():
                flash("Username already exists", "danger")
            else:
                user = User(
                    username=username,
                    full_name=full_name,
                    role=role,
                    must_change_password=True,
                )
                user.set_password(password)
                db.session.add(user)
                db.session.commit()
                flash(f"User {full_name} created", "success")

        elif action == "delete":
            user_id = request.form.get("user_id")
            user = db.session.get(User, int(user_id))
            if user and user.id != session.get("user_id"):
                username = user.username
                db.session.delete(user)
                db.session.commit()
                flash(f"User {username} deleted", "success")
            elif user and user.id == session.get("user_id"):
                flash("Cannot delete your own account", "danger")

        elif action == "reset_password":
            user_id = request.form.get("user_id")
            new_password = request.form.get("new_password", "")

            if not new_password or len(new_password) < 6:
                flash("Password must be at least 6 characters", "danger")
            else:
                user = db.session.get(User, int(user_id))
                user.set_password(new_password)
                user.must_change_password = True
                db.session.commit()
                flash(f"Password reset for {user.username}", "success")

    users = User.query.all()
    return render_template("users.html", users=users, display_codes=display_codes)


@app.route("/dashboard")
@login_required
def dashboard():
    date_range = request.args.get("date_range", "this_month")
    party_id = request.args.get("party")
    search = request.args.get("search", "").strip()
    sort_by = request.args.get("sort_by", "revenue")

    # Get date range
    period_start, period_end = get_date_range(date_range)

    # Get last month dates for calculations
    last_month_start, last_month_end = get_last_month_dates(period_start)

    # Get additional date ranges for calculations
    two_months_back_start, two_months_back_end = get_two_months_back_dates(
        last_month_start
    )
    three_months_start, three_months_end = get_three_months_back_dates(
        two_months_back_start
    )
    four_months_start, four_months_end = get_four_months_back_dates(three_months_start)
    last_3m_start, last_3m_end = get_last_3_months_dates(last_month_end)
    this_month_last_year_start, this_month_last_year_end = (
        get_this_month_last_year_dates()
    )

    this_month_last_year_label = (
        f"{get_month_name(datetime.now().month)} {datetime.now().year - 1}"
    )

    # Query invoices with filters
    invoices, parties = get_filtered_invoices(period_start, period_end, party_id)

    # Calculate GST once per invoice to avoid repeated calculations
    invoice_gst_data = [(inv, inv.calculate_gst()) for inv in invoices]

    # Calculate revenue and GST
    this_month_revenue, this_month_gst, invoice_count, unlocked_count = (
        calculate_revenue_and_gst(invoice_gst_data)
    )

    # Pre-filter invoices by date ranges to minimize queries
    (
        this_month_invs_all,
        last_month_invs_all,
        two_months_back_invs_all,
        three_months_back_invs_all,
        four_months_back_invs_all,
        last_3m_invs_all,
        this_month_last_year_invs_all,
    ) = get_prefiltered_invoices(
        period_start,
        period_end,
        last_month_start,
        last_month_end,
        two_months_back_start,
        two_months_back_end,
        three_months_start,
        three_months_end,
        four_months_start,
        four_months_end,
        last_3m_start,
        last_3m_end,
        this_month_last_year_start,
        this_month_last_year_end,
    )

    # Pre-calculate GST data for all filtered invoices
    (
        this_month_invs_with_gst,
        last_month_invs_with_gst,
        two_months_back_invs_with_gst,
        three_months_back_invs_with_gst,
        four_months_back_invs_with_gst,
        last_3m_invs_with_gst,
        this_month_last_year_invs_with_gst,
    ) = precalculate_gst_data(
        this_month_invs_all,
        last_month_invs_all,
        two_months_back_invs_all,
        three_months_back_invs_all,
        four_months_back_invs_all,
        last_3m_invs_all,
        this_month_last_year_invs_all,
    )

    # Calculate party growth data
    party_growth_data, max_revenue = calculate_party_growth_data(
        parties,
        this_month_invs_all,
        this_month_invs_with_gst,
        last_month_invs_all,
        last_month_invs_with_gst,
        two_months_back_invs_all,
        two_months_back_invs_with_gst,
        three_months_back_invs_all,
        three_months_back_invs_with_gst,
        four_months_back_invs_all,
        four_months_back_invs_with_gst,
        last_3m_invs_all,
        last_3m_invs_with_gst,
        this_month_last_year_invs_all,
        this_month_last_year_invs_with_gst,
    )

    # Sort party growth data
    party_growth_data = sort_party_growth_data(party_growth_data, sort_by)

    # Show ALL parties (not just those with recent invoices)
    party_growth_data = add_missing_parties(party_growth_data, parties)

    # Overall revenue change calculation
    revenue_change = calculate_revenue_change(this_month_revenue, party_growth_data)

    # Pending invoices
    pending_count, pending_amount = get_pending_invoices_info()

    # Chart data
    month_names = [
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
        "Jan",
        "Feb",
        "Mar",
    ]
    chart_labels, chart_data = prepare_chart_data(month_names)

    # Party stats for top parties
    this_month_party_stats, top_parties = calculate_top_parties(invoices)

    # Miscellaneous data
    company_name = Settings.get("company_name", "")
    greeting = get_greeting()
    trend_months = get_trend_months()

    return render_template(
        "dashboard.html",
        invoices=invoices[:5],
        parties=parties,
        pending_count=pending_count,
        pending_amount=pending_amount,
        this_month_revenue=this_month_revenue,
        this_month_gst=this_month_gst,
        invoice_count=invoice_count,
        revenue_change=revenue_change,
        chart_labels=chart_labels,
        chart_data=chart_data,
        top_parties=top_parties,
        date_range=date_range,
        selected_party=party_id,
        greeting=greeting,
        company_name=company_name,
        party_growth_data=party_growth_data,
        max_revenue=max_revenue,
        sort_by=sort_by,
        trend_months=trend_months,
        current_year=datetime.now().year,
        this_month_last_year_label=this_month_last_year_label,
    )
    this_month_gst = sum(
        (gst_data.get("cgst", 0) or 0)
        + (gst_data.get("sgst", 0) or 0)
        + (gst_data.get("igst", 0) or 0)
        for inv, gst_data in invoice_gst_data
        if not inv.is_rcm
    )
    invoice_count = len(invoices)
    unlocked_count = sum(1 for inv, _ in invoice_gst_data if not inv.locked)

    # Last month dates (previous month)
    last_month_start = (period_start - timedelta(days=1)).replace(day=1)
    last_month_end = period_start - timedelta(days=1)
    last_month_invoices = Invoice.query.filter(
        Invoice.invoice_date >= last_month_start, Invoice.invoice_date <= last_month_end
    ).all()

    # Previous to previous month (2 months ago)
    two_months_back_start = (last_month_start - timedelta(days=1)).replace(day=1)
    two_months_back_end = last_month_start - timedelta(days=1)

    # Three months back
    three_months_start = (two_months_back_start - timedelta(days=1)).replace(day=1)
    three_months_end = two_months_back_start - timedelta(days=1)

    # Four months back
    four_months_start = (three_months_start - timedelta(days=1)).replace(day=1)
    four_months_end = three_months_start - timedelta(days=1)

    # Pre-filter invoices by date ranges to minimize queries
    this_month_invs_all = Invoice.query.filter(
        Invoice.invoice_date >= period_start,
        Invoice.invoice_date <= period_end,
    ).all()

    last_month_invs_all = Invoice.query.filter(
        Invoice.invoice_date >= last_month_start,
        Invoice.invoice_date <= last_month_end,
    ).all()

    two_months_back_invs_all = Invoice.query.filter(
        Invoice.invoice_date >= two_months_back_start,
        Invoice.invoice_date <= two_months_back_end,
    ).all()

    three_months_back_invs_all = Invoice.query.filter(
        Invoice.invoice_date >= three_months_start,
        Invoice.invoice_date <= three_months_end,
    ).all()

    four_months_back_invs_all = Invoice.query.filter(
        Invoice.invoice_date >= four_months_start,
        Invoice.invoice_date <= four_months_end,
    ).all()

    last_3m_invs_all = Invoice.query.filter(
        Invoice.invoice_date >= last_3m_start,
        Invoice.invoice_date <= last_3m_end,
    ).all()

    this_month_last_year_invs_all = Invoice.query.filter(
        Invoice.invoice_date >= this_month_last_year_start,
        Invoice.invoice_date <= this_month_last_year_end,
    ).all()

    # Pre-calculate GST data for all filtered invoices
    this_month_invs_with_gst = [
        (inv, inv.calculate_gst()) for inv in this_month_invs_all
    ]
    last_month_invs_with_gst = [
        (inv, inv.calculate_gst()) for inv in last_month_invs_all
    ]
    two_months_back_invs_with_gst = [
        (inv, inv.calculate_gst()) for inv in two_months_back_invs_all
    ]
    three_months_back_invs_with_gst = [
        (inv, inv.calculate_gst()) for inv in three_months_back_invs_all
    ]
    four_months_back_invs_with_gst = [
        (inv, inv.calculate_gst()) for inv in four_months_back_invs_all
    ]
    last_3m_invs_with_gst = [(inv, inv.calculate_gst()) for inv in last_3m_invs_all]
    this_month_last_year_invs_with_gst = [
        (inv, inv.calculate_gst()) for inv in this_month_last_year_invs_all
    ]

    for party in all_parties:
        # This month revenue
        this_month_invs = [
            inv for inv in this_month_invs_all if inv.party_id == party.id
        ]
        this_month_rev = sum(
            gst_data.get("subtotal", 0) or 0
            for inv, gst_data in this_month_invs_with_gst
            if inv.party_id == party.id
        )

        # This month last year
        this_m_last_yr_invs = [
            inv for inv in this_month_last_year_invs_all if inv.party_id == party.id
        ]
        this_month_last_year_rev = sum(
            gst_data.get("subtotal", 0) or 0
            for inv, gst_data in this_month_last_year_invs_with_gst
            if inv.party_id == party.id
        )

        # Last month revenue
        last_m_invs = [inv for inv in last_month_invs_all if inv.party_id == party.id]
        last_month_rev = sum(
            gst_data.get("subtotal", 0) or 0
            for inv, gst_data in last_month_invs_with_gst
            if inv.party_id == party.id
        )

        # Previous to previous month (2 months ago)
        two_m_invs = [
            inv for inv in two_months_back_invs_all if inv.party_id == party.id
        ]
        two_months_ago_rev = sum(
            gst_data.get("subtotal", 0) or 0
            for inv, gst_data in two_months_back_invs_with_gst
            if inv.party_id == party.id
        )

        # Three months ago
        three_m_invs = [
            inv for inv in three_months_back_invs_all if inv.party_id == party.id
        ]
        three_months_ago_rev = sum(
            gst_data.get("subtotal", 0) or 0
            for inv, gst_data in three_months_back_invs_with_gst
            if inv.party_id == party.id
        )

        # Four months ago
        four_m_invs = [
            inv for inv in four_months_back_invs_all if inv.party_id == party.id
        ]
        four_months_ago_rev = sum(
            gst_data.get("subtotal", 0) or 0
            for inv, gst_data in four_months_back_invs_with_gst
            if inv.party_id == party.id
        )

        # Last 3 months revenue
        last_3m_invs = [inv for inv in last_3m_invs_all if inv.party_id == party.id]
        last_3m_rev = sum(
            gst_data.get("subtotal", 0) or 0
            for inv, gst_data in last_3m_invs_with_gst
            if inv.party_id == party.id
        )

        # Last 6 months for trend (monthly breakdown)
        trend_data = []
        for i in range(5, -1, -1):
            trg_month = (now.month - i + 12) % 12 + 1
            trg_year = now.year if (now.month - i) > 0 else now.year - 1
            if now.month - i <= 0:
                trg_year = now.year - 1
            month_invoices = Invoice.query.filter(
                Invoice.party_id == party.id,
                db.extract("month", Invoice.invoice_date) == trg_month,
                db.extract("year", Invoice.invoice_date) == trg_year,
            ).all()

            # Pre-calculate GST data for trend month invoices
            month_invoices_with_gst = [
                (inv, inv.calculate_gst()) for inv in month_invoices
            ]
            month_rev = sum(
                gst_data.get("subtotal", 0) or 0
                for _, gst_data in month_invoices_with_gst
            )
            trend_data.append(month_rev)

        # Calculate growth %
        if last_month_rev > 0:
            growth_pct = ((this_month_rev - last_month_rev) / last_month_rev) * 100
        else:
            growth_pct = 0 if this_month_rev == 0 else 100

        if this_month_rev > max_revenue:
            max_revenue = this_month_rev

        if (
            this_month_rev > 0
            or this_month_last_year_rev > 0
            or last_month_rev > 0
            or two_months_ago_rev > 0
            or three_months_ago_rev > 0
            or four_months_ago_rev > 0
            or last_3m_rev > 0
        ):
            party_growth_data.append(
                {
                    "name": party.name,
                    "this_month": this_month_rev,
                    "this_month_last_year": this_month_last_year_rev,
                    "last_month": last_month_rev,
                    "two_months_ago": two_months_ago_rev,
                    "three_months_ago": three_months_ago_rev,
                    "four_months_ago": four_months_ago_rev,
                    "last_3m": last_3m_rev,
                    "growth": growth_pct,
                    "trend": trend_data,
                }
            )

    # Sort by revenue (default) or by name
    if sort_by == "name":
        party_growth_data = sorted(
            party_growth_data, key=lambda x: x["name"], reverse=False
        )
    else:
        party_growth_data = sorted(
            party_growth_data, key=lambda x: x["this_month"], reverse=True
        )

    # Show ALL parties (not just those with recent invoices)
    all_parties = Party.query.all()
    party_dict = {p.name: p for p in all_parties}
    for party in all_parties:
        party_name = party.name
        existing = next((x for x in party_growth_data if x["name"] == party_name), None)
        if not existing:
            party_growth_data.append(
                {
                    "name": party.name,
                    "this_month": 0,
                    "this_month_last_year": 0,
                    "last_month": 0,
                    "two_months_ago": 0,
                    "three_months_ago": 0,
                    "four_months_ago": 0,
                    "last_3m": 0,
                    "growth": 0,
                    "trend": [0, 0, 0, 0, 0, 0],
                }
            )
    if sort_by == "name":
        party_growth_data = sorted(
            party_growth_data, key=lambda x: x["name"], reverse=False
        )

    # Overall revenue change calculation
    total_last_month = sum(p["last_month"] for p in party_growth_data)
    if total_last_month > 0:
        revenue_change = (
            (this_month_revenue - total_last_month) / total_last_month
        ) * 100
    else:
        revenue_change = 0

    pending_invoices = Invoice.query.filter(Invoice.invoice_no.is_(None)).all()
    pending_count = len(pending_invoices)
    pending_amount = sum(
        inv.calculate_gst().get("total", 0) or 0 for inv in pending_invoices
    )

    months = [
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
        "Jan",
        "Feb",
        "Mar",
    ]

    chart_labels = []
    chart_data = []

    month_names = [
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
        "Jan",
        "Feb",
        "Mar",
    ]

    for i in range(5, -1, -1):
        m = now.month - i
        y = now.year
        if m < 1:
            m += 12
            y -= 1

        start = datetime(y, m, 1).date()
        if m == 12:
            end = datetime(y + 1, 1, 1).date()
        else:
            end = datetime(y, m + 1, 1).date()

        month_invoices = Invoice.query.filter(
            Invoice.invoice_date >= start, Invoice.invoice_date < end
        ).all()

        # Pre-calculate GST data for month invoices
        month_invoices_with_gst = [(inv, inv.calculate_gst()) for inv in month_invoices]

        revenue = sum(
            gst_data.get("subtotal", 0) or 0 for _, gst_data in month_invoices_with_gst
        )
        chart_labels.append(month_names[(m - 4 + 12) % 12])
        chart_data.append(revenue)

    this_month_party_stats = {}
    for inv in invoices:
        party_name = inv.party.name if inv.party else "Unknown"
        if party_name not in this_month_party_stats:
            this_month_party_stats[party_name] = {"count": 0, "revenue": 0}
        this_month_party_stats[party_name]["count"] += 1
        this_month_party_stats[party_name]["revenue"] += (
            inv.calculate_gst().get("subtotal", 0) or 0
        )

    top_parties = sorted(
        this_month_party_stats.items(), key=lambda x: x[1]["revenue"], reverse=True
    )[:5]

    company_name = Settings.get("company_name", "")
    greeting = get_greeting()

    return render_template(
        "dashboard.html",
        invoices=invoices[:5],
        parties=parties,
        pending_count=pending_count,
        pending_amount=pending_amount,
        this_month_revenue=this_month_revenue,
        this_month_gst=this_month_gst,
        invoice_count=invoice_count,
        revenue_change=revenue_change,
        chart_labels=chart_labels,
        chart_data=chart_data,
        top_parties=top_parties,
        date_range=date_range,
        selected_party=party_id,
        greeting=greeting,
        company_name=company_name,
        party_growth_data=party_growth_data,
        max_revenue=max_revenue,
        sort_by=sort_by,
        trend_months=trend_months,
        current_year=current_year,
        this_month_last_year_label=this_month_last_year_label,
    )


@app.route("/party/api/<int:party_id>")
def party_api(party_id):
    party = Party.query.get_or_404(party_id)
    return jsonify(
        {
            "name": party.name,
            "gstin": party.gstin,
            "pan": party.pan,
            "address": party.address,
            "state": party.state,
            "state_code": party.state_code,
        }
    )


@app.route("/parties", methods=["GET", "POST"])
@login_required
def parties():
    # Check authentication
    if not session.get("user_id"):
        return redirect(url_for("login"))

    if request.method == "POST":
        name = request.form.get("name")
        gstin = request.form.get("gstin")
        amc_code = request.form.get("amc_code")
        pan = request.form.get("pan")
        address = request.form.get("address")
        state = request.form.get("state")
        state_code = request.form.get("state_code")
        email = request.form.get("email")
        phone = request.form.get("phone")

        existing = Party.query.filter_by(gstin=gstin).first()
        if existing:
            flash(f"Party with GSTIN {gstin} already exists", "danger")
            return redirect(url_for("parties"))

        party = Party(
            name=name,
            gstin=gstin,
            amc_code=amc_code,
            pan=pan,
            address=address,
            state=state,
            state_code=state_code,
            email=email,
            phone=phone,
        )
        db.session.add(party)
        db.session.commit()
        flash("Party added successfully", "success")
        return redirect(url_for("parties"))

    sort_by = request.args.get("sort_by", "name")
    sort_dir = request.args.get("sort_dir", "asc")

    sort_column = getattr(Party, sort_by, Party.name)
    if sort_dir == "desc":
        query = Party.query.order_by(sort_column.desc())
    else:
        query = Party.query.order_by(sort_column.asc())

    parties = query.all()
    return render_template(
        "parties.html", parties=parties, sort_by=sort_by, sort_dir=sort_dir
    )


@app.route("/export/parties")
def export_parties():
    parties = Party.query.all()
    return export_parties_response(parties)


@app.route("/export/parties/selected", methods=["POST"])
def export_selected_parties():
    ids_raw = request.form.get("party_ids", "")
    party_ids = [int(x) for x in ids_raw.split(",") if x] if ids_raw else []
    if not party_ids:
        flash("No parties selected", "warning")
        return redirect(url_for("parties"))
    parties = Party.query.filter(Party.id.in_(party_ids)).all()
    return export_parties_response(parties)


def export_parties_response(parties):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "name",
            "gstin",
            "pan",
            "amc_code",
            "address",
            "state",
            "state_code",
            "email",
            "phone",
        ]
    )

    for party in parties:
        writer.writerow(
            [
                party.name,
                party.gstin,
                party.pan or "",
                party.amc_code or "",
                party.address or "",
                party.state or "",
                party.state_code or "",
                party.email or "",
                party.phone or "",
            ]
        )

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"parties_export_{datetime.now().strftime('%Y%m%d')}.csv",
    )


@app.route("/party/edit/<int:party_id>", methods=["GET", "POST"])
def edit_party(party_id):
    party = Party.query.get_or_404(party_id)

    if request.method == "POST":
        party.name = request.form.get("name")
        party.gstin = request.form.get("gstin")
        party.pan = request.form.get("pan")
        party.amc_code = request.form.get("amc_code")
        party.address = request.form.get("address")
        party.state = request.form.get("state")
        party.state_code = request.form.get("state_code")
        party.email = request.form.get("email")
        party.phone = request.form.get("phone")

        existing = Party.query.filter(
            Party.gstin == party.gstin, Party.id != party_id
        ).first()
        if existing:
            flash(f"Party with GSTIN {party.gstin} already exists", "danger")
            return redirect(url_for("edit_party", party_id=party_id))

        db.session.commit()
        flash("Party updated successfully", "success")
        return redirect(url_for("parties"))

    return render_template("edit_party.html", party=party)


@app.route("/party/delete/<int:party_id>")
def delete_party(party_id):
    party = Party.query.get_or_404(party_id)

    invoice_count = Invoice.query.filter_by(party_id=party_id).count()
    if invoice_count > 0:
        flash(
            f"Cannot delete party with {invoice_count} existing invoice(s). Delete invoices first.",
            "danger",
        )
        return redirect(url_for("parties"))

    db.session.delete(party)
    db.session.commit()
    flash("Party deleted successfully", "success")
    return redirect(url_for("parties"))


@app.route("/invoices", methods=["GET"])
def manage_invoices():
    # Check authentication
    if not session.get("user_id"):
        return redirect(url_for("login"))

    year = request.args.get("year")
    month = request.args.get("month")
    search = request.args.get("search", "").strip()
    sort_by = request.args.get("sort_by", "date")
    sort_dir = request.args.get("sort_dir", "desc")
    selected_year = year
    selected_month = month
    search_query = search
    tax_type_filter = request.args.get("tax_type", "")
    party_filter = request.args.get("party", "")
    status_filter = request.args.get("status", "")

    query = Invoice.query.join(Party)

    if year:
        query = query.filter(db.extract("year", Invoice.invoice_date) == int(year))
    if month:
        query = query.filter(db.extract("month", Invoice.invoice_date) == int(month))
    if search:
        query = query.filter(
            db.or_(
                Invoice.invoice_no.ilike(f"%{search}%"),
                Invoice.reference_serial_no.ilike(f"%{search}%"),
                Party.gstin.ilike(f"%{search}%"),
                Party.name.ilike(f"%{search}%"),
            )
        )
    if tax_type_filter:
        query = query.filter(Invoice.tax_type == tax_type_filter)
    if party_filter:
        query = query.filter(Invoice.party_id == int(party_filter))
    if status_filter == "pending":
        query = query.filter(Invoice.invoice_no.is_(None))
    elif status_filter == "completed":
        query = query.filter(Invoice.invoice_no != None)

    sort_column = Invoice.invoice_no
    if sort_by == "date":
        sort_column = Invoice.invoice_date
    elif sort_by == "party":
        sort_column = Party.name

    if sort_dir == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    invoices = query.all()
    parties = Party.query.all()
    pending_count = Invoice.query.filter(Invoice.invoice_no.is_(None)).count()

    all_invoices = Invoice.query.all()
    years = sorted(
        set(inv.invoice_date.year for inv in all_invoices if inv.invoice_date)
    )
    if not years:
        years = [datetime.now().year]

    return render_template(
        "invoice_management.html",
        invoices=invoices,
        parties=parties,
        selected_year=selected_year,
        selected_month=selected_month,
        search_query=search_query,
        years=years,
        sort_by=sort_by,
        sort_dir=sort_dir,
        tax_type_filter=tax_type_filter,
        party_filter=party_filter,
        status_filter=status_filter,
        pending_count=pending_count,
    )


@app.route("/invoice/create", methods=["GET", "POST"])
def create_invoice():
    parties = Party.query.all()

    if request.method == "POST":
        party_id = request.form.get("party_id")
        reference_serial_no = request.form.get("reference_serial_no", "").strip()
        sac_hsn_code = request.form.get("sac_hsn_code", "").strip()

        if not party_id:
            flash("Party selection is required.", "danger")
            return redirect(url_for("create_invoice"))

        if not sac_hsn_code:
            flash("SAC/HSN code is required.", "danger")
            return redirect(url_for("create_invoice"))

        if reference_serial_no:
            existing = Invoice.query.filter(
                Invoice.reference_serial_no == reference_serial_no
            ).first()
            if existing:
                flash(
                    f"Reference serial number {reference_serial_no} already exists",
                    "danger",
                )
                return redirect(url_for("create_invoice"))

        invoice_date = datetime.strptime(
            request.form.get("invoice_date"), "%Y-%m-%d"
        ).date()
        tax_type = request.form.get("tax_type")
        place_of_supply = request.form.get("place_of_supply")
        sac_hsn_code = request.form.get("sac_hsn_code")
        reverse_charge = float(request.form.get("reverse_charge") or 0)
        is_rcm = request.form.get("is_rcm") == "on"
        distributor_code = request.form.get("distributor_code", "").strip()

        if is_rcm and reverse_charge <= 0:
            flash("Reverse Charge amount is required when RCM is enabled.", "danger")
            return redirect(url_for("create_invoice"))

        invoice = Invoice(
            reference_serial_no=reference_serial_no,
            invoice_date=invoice_date,
            party_id=party_id,
            tax_type=tax_type,
            place_of_supply=place_of_supply,
            sac_hsn_code=sac_hsn_code,
            reverse_charge=reverse_charge,
            is_rcm=is_rcm,
            distributor_code=distributor_code,
        )

        # Local copy of party details
        party = db.session.get(Party, party_id)
        if party:
            invoice.party_name = party.name
            invoice.party_address = party.address
            invoice.party_gstin = party.gstin
            invoice.party_pan = party.pan
            invoice.party_state = party.state
            invoice.party_state_code = party.state_code

        db.session.add(invoice)
        db.session.flush()

        items_desc = request.form.getlist("item_description[]")
        items_taxable = request.form.getlist("item_taxable_value[]")
        items_cgst_rate = request.form.getlist("item_cgst_rate[]")
        items_sgst_rate = request.form.getlist("item_sgst_rate[]")
        items_igst_rate = request.form.getlist("item_igst_rate[]")

        for i in range(len(items_desc)):
            if items_desc[i].strip():
                taxable = float(items_taxable[i]) if i < len(items_taxable) else 0
                cgst_rate = float(items_cgst_rate[i]) if i < len(items_cgst_rate) else 0
                sgst_rate = float(items_sgst_rate[i]) if i < len(items_sgst_rate) else 0
                igst_rate = float(items_igst_rate[i]) if i < len(items_igst_rate) else 0

                is_valid, error_msg = validate_tax_rates(
                    tax_type, is_rcm, cgst_rate, sgst_rate, igst_rate
                )
                if not is_valid:
                    flash(f"Item '{items_desc[i][:30]}': {error_msg}", "danger")
                    return redirect(url_for("create_invoice"))

                cgst_amt = taxable * cgst_rate / 100 if tax_type == "INTRA" else 0
                sgst_amt = taxable * sgst_rate / 100 if tax_type == "INTRA" else 0
                igst_amt = taxable * igst_rate / 100 if tax_type == "INTER" else 0

                item = InvoiceItem(
                    invoice_id=invoice.id,
                    description=items_desc[i],
                    taxable_value=taxable,
                    cgst_rate=cgst_rate,
                    cgst_amt=cgst_amt,
                    sgst_rate=sgst_rate,
                    sgst_amt=sgst_amt,
                    igst_rate=igst_rate,
                    igst_amt=igst_amt,
                )
                db.session.add(item)

        db.session.commit()
        flash("Invoice created successfully", "success")
        return redirect(url_for("manage_invoices"))

    return render_template("create_invoice.html", parties=parties)


@app.route("/invoice/<int:invoice_id>")
def view_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    return render_template(
        "invoice_preview.html",
        invoice=invoice,
        settings={
            "company_name": Settings.get("company_name", ""),
            "address": Settings.get("address", ""),
            "gstin": Settings.get("gstin", ""),
            "pan": Settings.get("pan", ""),
            "logo": Settings.get("logo", ""),
        },
    )


@app.route("/invoice/delete/<int:invoice_id>")
def delete_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    db.session.delete(invoice)
    db.session.commit()
    flash("Invoice deleted successfully", "success")
    return redirect(url_for("dashboard"))


@app.route("/invoice/pdf/<int:invoice_id>")
def generate_pdf(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)

    html_content = render_template(
        "invoice_pdf_template.html",
        invoice=invoice,
        settings={
            "company_name": Settings.get("company_name", ""),
            "address": Settings.get("address", ""),
            "gstin": Settings.get("gstin", ""),
            "pan": Settings.get("pan", ""),
            "logo": Settings.get("logo", ""),
            "place_of_supply": Settings.get("place_of_supply", ""),
            "state_code": Settings.get("state_code", ""),
        },
    )

    pdf_path = get_export_path(invoice)
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

    HTML(string=html_content).write_pdf(pdf_path)

    download_name = (
        f"{invoice.invoice_no}.pdf"
        if invoice.invoice_no
        else f"invoice_{invoice.id}.pdf"
    )
    return send_file(pdf_path, as_attachment=True, download_name=download_name)


@app.route("/invoice/preview-html/<int:invoice_id>")
def preview_html(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    return render_template("invoice_print.html", invoice=invoice)


@app.route("/invoice/generate-numbers", methods=["POST"])
def generate_numbers():
    count = generate_invoice_numbers()
    if count > 0:
        flash(f"Generated invoice numbers for {count} invoices", "success")
    else:
        flash("No pending invoices to generate numbers for", "warning")
    return redirect(url_for("manage_invoices"))


@app.route("/batch/export", methods=["POST"])
def batch_export():
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = [x.strip() for x in ids_raw.split(",") if x.strip()]

    if not invoice_ids:
        flash("No invoices selected", "warning")
        return redirect(url_for("dashboard"))

    zip_path = os.path.join(
        app.config["EXPORT_FOLDER"],
        f"invoices_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
    )

    exported_count = 0
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for invoice_id in invoice_ids:
            invoice = db.session.get(Invoice, int(invoice_id))
            if not invoice or not invoice.invoice_no or not invoice.locked:
                continue

            html_content = render_template(
                "invoice_pdf_template.html",
                invoice=invoice,
                settings={
                    "company_name": Settings.get("company_name", ""),
                    "arn_number": Settings.get("arn_number", ""),
                    "address": Settings.get("address", ""),
                    "gstin": Settings.get("gstin", ""),
                    "pan": Settings.get("pan", ""),
                    "logo": Settings.get("logo", ""),
                    "place_of_supply": Settings.get("place_of_supply", ""),
                    "state_code": Settings.get("state_code", ""),
                },
            )
            pdf_path = get_export_path(invoice)
            os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

            HTML(string=html_content).write_pdf(pdf_path)

            zip_filename = os.path.basename(pdf_path)
            zip_file.write(pdf_path, zip_filename)
            os.remove(pdf_path)
            exported_count += 1

    if exported_count == 0:
        flash("No locked invoices with invoice numbers to export", "warning")
        return redirect(url_for("manage_invoices"))

    return send_file(zip_path, as_attachment=True, download_name=f"invoices_batch.zip")


@app.route("/batch/export/excel", methods=["POST"])
def batch_export_excel():
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = [x.strip() for x in ids_raw.split(",") if x.strip()]

    if not invoice_ids:
        flash("No invoices selected", "warning")
        return redirect(url_for("manage_invoices"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    header_fill = PatternFill(
        start_color="4F46E5", end_color="4F46E5", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    headers = ["Invoice Date", "Invoice Number", "GSTIN", "Party Name", "Taxable Value"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    exported_count = 0
    for invoice_id in invoice_ids:
        invoice = db.session.get(Invoice, int(invoice_id))
        if not invoice or not invoice.invoice_no or not invoice.locked:
            continue

        gst_data = invoice.calculate_gst()
        taxable = gst_data.get("subtotal", 0) or 0

        ws.cell(
            row=row_idx,
            column=1,
            value=invoice.invoice_date.strftime("%d-%m-%Y")
            if invoice.invoice_date
            else "",
        )
        ws.cell(row=row_idx, column=2, value=invoice.invoice_no)
        ws.cell(
            row=row_idx, column=3, value=invoice.party.gstin if invoice.party else ""
        )
        ws.cell(
            row=row_idx, column=4, value=invoice.party.name if invoice.party else ""
        )
        ws.cell(row=row_idx, column=5, value=taxable)
        row_idx += 1
        exported_count += 1

    if exported_count == 0:
        flash("No locked invoices with invoice numbers to export", "warning")
        return redirect(url_for("manage_invoices"))

    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"invoices_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return make_response(
        send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    )


@app.route("/import/parties", methods=["GET", "POST"])
def import_parties():
    if request.method == "POST":
        file = request.files.get("csv_file")

        if not file or file.filename == "":
            flash("No file selected", "danger")
            return redirect(url_for("import_parties"))

        if not file.filename.endswith(".csv"):
            flash("Please upload a CSV file", "danger")
            return redirect(url_for("import_parties"))

        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)

        imported = 0
        errors = []

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                gstin = (row.get("gstin") or "").strip().upper()
                name = (row.get("name") or "").strip()
                provided_pan = (row.get("pan") or "").strip().upper()
                amc_code = (row.get("amc_code") or "").strip()
                address = (row.get("address") or "").strip()
                state = (row.get("state") or "").strip()
                state_code = (row.get("state_code") or "").strip()

                if not name or not gstin:
                    errors.append(f"Row {row_num}: Missing name or GSTIN")
                    continue

                pan = provided_pan if provided_pan else extract_pan_from_gstin(gstin)

                existing = Party.query.filter_by(gstin=gstin).first()
                if existing:
                    existing.name = name
                    existing.pan = pan
                    if amc_code:
                        existing.amc_code = amc_code
                    if address:
                        existing.address = address
                    if state:
                        existing.state = state
                    if state_code:
                        existing.state_code = state_code
                    imported += 1
                else:
                    party = Party(
                        name=name,
                        gstin=gstin,
                        pan=pan,
                        amc_code=amc_code,
                        address=address,
                        state=state,
                        state_code=state_code,
                    )
                    db.session.add(party)
                    imported += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        if imported > 0:
            db.session.commit()
            flash(f"Successfully imported {imported} parties", "success")

        if errors:
            for error in errors[:10]:
                flash(error, "warning")

        return redirect(url_for("parties"))

    return render_template("import_parties.html")


@app.route("/import/invoices", methods=["GET", "POST"])
def import_invoices():
    if request.method == "POST":
        file = request.files.get("csv_file")

        if not file or file.filename == "":
            flash("No file selected", "danger")
            return redirect(url_for("import_invoices"))

        if not file.filename.endswith(".csv"):
            flash("Please upload a CSV file", "danger")
            return redirect(url_for("import_invoices"))

        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)

        imported = 0
        updated = 0
        errors = []

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                invoice_no = (row.get("invoice_no") or "").strip()
                reference_serial_no = (row.get("reference_serial_no") or "").strip()
                invoice_date_str = (row.get("invoice_date") or "").strip()
                party_gstin = (row.get("party_gstin") or "").strip().upper()
                description = (row.get("description") or "").strip()
                sac_hsn_code = (row.get("sac_hsn_code") or "").strip()
                taxable_value = parse_number(row.get("taxable_value"))
                tax_type = parse_tax_type(row.get("tax_type"))
                cgst_rate = parse_percentage(row.get("cgst_rate"))
                sgst_rate = parse_percentage(row.get("sgst_rate"))
                igst_rate = parse_percentage(row.get("igst_rate"))
                place_of_supply = (row.get("place_of_supply") or "").strip()
                reverse_charge = parse_number(row.get("reverse_charge"))
                is_rcm = (row.get("is_rcm") or "").strip() == "1"
                distributor_code = (row.get("distributor_code") or "").strip()

                if not party_gstin or not invoice_date_str:
                    errors.append(f"Row {row_num}: Missing GSTIN or Invoice Date")
                    continue

                party = Party.query.filter_by(gstin=party_gstin).first()
                if not party:
                    errors.append(
                        f"Row {row_num}: Party with GSTIN {party_gstin} not found"
                    )
                    continue

                invoice_date = parse_date(invoice_date_str)
                if not invoice_date:
                    errors.append(
                        f"Row {row_num}: Invalid date format - {invoice_date_str}"
                    )
                    continue

                if is_rcm and reverse_charge <= 0:
                    errors.append(
                        f"Row {row_num}: Reverse Charge amount is required when RCM is enabled"
                    )
                    continue

                invoice = None
                if invoice_no:
                    invoice = Invoice.query.filter_by(invoice_no=invoice_no).first()
                elif reference_serial_no:
                    invoice = Invoice.query.filter_by(
                        reference_serial_no=reference_serial_no
                    ).first()

                if invoice:
                    if invoice.locked:
                        errors.append(
                            f"Row {row_num}: Invoice {invoice.invoice_no or invoice.id} is locked and cannot be updated"
                        )
                        continue

                    invoice.invoice_no = invoice_no or invoice.invoice_no
                    invoice.reference_serial_no = (
                        reference_serial_no or invoice.reference_serial_no
                    )
                    invoice.invoice_date = invoice_date
                    invoice.party_id = party.id
                    invoice.tax_type = tax_type
                    invoice.place_of_supply = place_of_supply
                    invoice.sac_hsn_code = sac_hsn_code
                    invoice.reverse_charge = reverse_charge
                    invoice.is_rcm = is_rcm
                    invoice.distributor_code = distributor_code

                    # Update local copy
                    invoice.party_name = party.name
                    invoice.party_address = party.address
                    invoice.party_gstin = party.gstin
                    invoice.party_pan = party.pan
                    invoice.party_state = party.state
                    invoice.party_state_code = party.state_code
                    updated += 1
                else:
                    invoice = Invoice(
                        invoice_no=invoice_no or None,
                        reference_serial_no=reference_serial_no,
                        invoice_date=invoice_date,
                        party_id=party.id,
                        tax_type=tax_type,
                        place_of_supply=place_of_supply,
                        sac_hsn_code=sac_hsn_code,
                        reverse_charge=reverse_charge,
                        is_rcm=is_rcm,
                        distributor_code=distributor_code,
                    )
                    # Local copy
                    invoice.party_name = party.name
                    invoice.party_address = party.address
                    invoice.party_gstin = party.gstin
                    invoice.party_pan = party.pan
                    invoice.party_state = party.state
                    invoice.party_state_code = party.state_code
                    db.session.add(invoice)
                    db.session.flush()
                    imported += 1

                if invoice:
                    InvoiceItem.query.filter_by(invoice_id=invoice.id).delete()

                is_valid, error_msg = validate_tax_rates(
                    tax_type, is_rcm, cgst_rate, sgst_rate, igst_rate
                )
                if not is_valid:
                    errors.append(
                        f"Row {row_num}: Invalid tax configuration - {error_msg}"
                    )
                    continue

                cgst_amt = 0
                sgst_amt = 0
                igst_amt = 0
                if not is_rcm:
                    cgst_amt = (
                        taxable_value * cgst_rate / 100 if tax_type == "INTRA" else 0
                    )
                    sgst_amt = (
                        taxable_value * sgst_rate / 100 if tax_type == "INTRA" else 0
                    )
                    igst_amt = (
                        taxable_value * igst_rate / 100 if tax_type == "INTER" else 0
                    )

                item = InvoiceItem(
                    invoice_id=invoice.id,
                    description=description,
                    taxable_value=taxable_value,
                    cgst_rate=cgst_rate,
                    cgst_amt=cgst_amt,
                    sgst_rate=sgst_rate,
                    sgst_amt=sgst_amt,
                    igst_rate=igst_rate,
                    igst_amt=igst_amt,
                )
                db.session.add(item)

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        if imported > 0 or updated > 0:
            db.session.commit()
            flash(
                f"Successfully imported {imported} new and updated {updated} existing invoices",
                "success",
            )

        if errors:
            flash(
                f"Failed to import {len(errors)} rows. See below for first 10 errors:",
                "warning",
            )
            for error in errors[:10]:
                flash(error, "info")

        return redirect(url_for("manage_invoices"))

    return render_template("import_invoices.html")


@app.route("/invoice/sync-party/<int:invoice_id>")
def sync_party(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    if invoice.locked:
        return jsonify(
            {"success": False, "message": "Locked invoices cannot be synced"}
        ), 403
    party = db.session.get(Party, invoice.party_id)
    if not party:
        return jsonify({"success": False, "message": "Linked party not found"}), 404
    invoice.party_name = party.name
    invoice.party_address = party.address
    invoice.party_gstin = party.gstin
    invoice.party_pan = party.pan
    invoice.party_state = party.state
    invoice.party_state_code = party.state_code
    db.session.commit()
    return jsonify(
        {
            "success": True,
            "data": {
                "name": invoice.party_name,
                "address": invoice.party_address,
                "gstin": invoice.party_gstin,
                "pan": invoice.party_pan,
                "state": invoice.party_state,
                "state_code": invoice.party_state_code,
            },
        }
    )


@app.route("/invoice/batch-sync", methods=["POST"])
def batch_sync():
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = ids_raw.split(",") if ids_raw else []
    if not invoice_ids or invoice_ids == [""]:
        flash("No invoices selected", "warning")
        return redirect(url_for("manage_invoices"))
    synced_count, locked_count = 0, 0
    for inv_id in invoice_ids:
        inv = db.session.get(Invoice, int(inv_id))
        if inv:
            if inv.locked:
                locked_count += 1
                continue
            party = db.session.get(Party, inv.party_id)
            if party:
                inv.party_name = party.name
                inv.party_address = party.address
                inv.party_gstin = party.gstin
                inv.party_pan = party.pan
                inv.party_state = party.state
                inv.party_state_code = party.state_code
                synced_count += 1
    db.session.commit()
    if synced_count > 0:
        flash(
            f"Successfully synced party details for {synced_count} invoices", "success"
        )
    if locked_count > 0:
        flash(f"Skipped {locked_count} locked invoices", "info")
    return redirect(url_for("manage_invoices"))


@app.route("/invoice/batch-lock", methods=["POST"])
def batch_lock():
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = ids_raw.split(",") if ids_raw else []
    if not invoice_ids or invoice_ids == [""]:
        flash("No invoices selected", "warning")
        return redirect(url_for("manage_invoices"))
    locked_count = 0
    skipped_count = 0
    for inv_id in invoice_ids:
        inv = db.session.get(Invoice, int(inv_id))
        if inv:
            if inv.locked:
                skipped_count += 1
            elif not inv.invoice_no:
                skipped_count += 1
            else:
                inv.locked = True
                locked_count += 1
    db.session.commit()
    if locked_count > 0:
        flash(f"Successfully locked {locked_count} invoices", "success")
    if skipped_count > 0:
        flash(
            f"Skipped {skipped_count} invoice(s) - already locked or without invoice number",
            "info",
        )
    return redirect(url_for("manage_invoices"))


@app.route("/invoice/batch-unlock", methods=["POST"])
def batch_unlock():
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = ids_raw.split(",") if ids_raw else []
    if not invoice_ids or invoice_ids == [""]:
        flash("No invoices selected", "warning")
        return redirect(url_for("manage_invoices"))
    unlocked_count = 0
    skipped_count = 0
    for inv_id in invoice_ids:
        inv = db.session.get(Invoice, int(inv_id))
        if inv:
            if not inv.locked:
                skipped_count += 1
            else:
                inv.locked = False
                unlocked_count += 1
    db.session.commit()
    if unlocked_count > 0:
        flash(f"Successfully unlocked {unlocked_count} invoices", "success")
    if skipped_count > 0:
        flash(f"Skipped {skipped_count} invoice(s) - already unlocked", "info")
    return redirect(url_for("manage_invoices"))


@app.route("/invoice/batch-delete", methods=["POST"])
def batch_delete():
    ids_raw = request.form.get("invoice_ids", "")
    invoice_ids = ids_raw.split(",") if ids_raw else []
    if not invoice_ids or invoice_ids == [""]:
        flash("No invoices selected", "warning")
        return redirect(url_for("manage_invoices"))
    deleted_count = 0
    skipped_count = 0
    for inv_id in invoice_ids:
        inv = db.session.get(Invoice, int(inv_id))
        if inv and not inv.locked:
            if inv.credit_notes:
                skipped_count += 1
                continue
            db.session.delete(inv)
            deleted_count += 1
    db.session.commit()
    if deleted_count > 0:
        flash(f"Successfully deleted {deleted_count} invoices", "success")
    if skipped_count > 0:
        flash(f"Skipped {skipped_count} invoice(s) with credit notes", "warning")
    if deleted_count == 0 and skipped_count == 0:
        flash("No invoices deleted (some may be locked)", "warning")
    return redirect(url_for("manage_invoices"))


@app.route("/invoice/edit/<int:invoice_id>", methods=["GET", "POST"])
def edit_invoice(invoice_id):
    invoice = Invoice.query.get_or_404(invoice_id)
    if invoice.locked:
        flash("This invoice is locked and cannot be edited.", "danger")
        return redirect(url_for("manage_invoices"))
    parties = Party.query.all()

    if request.method == "POST":
        invoice_no = request.form.get("invoice_no", "").strip()
        reference_serial_no = request.form.get("reference_serial_no", "").strip()

        if invoice_no:
            existing = Invoice.query.filter(
                Invoice.invoice_no == invoice_no, Invoice.id != invoice_id
            ).first()
            if existing:
                flash(f"Invoice number {invoice_no} already exists", "danger")
                return redirect(url_for("edit_invoice", invoice_id=invoice_id))

        if reference_serial_no:
            existing = Invoice.query.filter(
                Invoice.reference_serial_no == reference_serial_no,
                Invoice.id != invoice_id,
            ).first()
            if existing:
                flash(
                    f"Reference serial number {reference_serial_no} already exists",
                    "danger",
                )
                return redirect(url_for("edit_invoice", invoice_id=invoice_id))

        invoice.invoice_no = invoice_no or None
        invoice.reference_serial_no = reference_serial_no or None
        invoice.party_id = request.form.get("party_id")
        invoice.invoice_date = datetime.strptime(
            request.form.get("invoice_date"), "%Y-%m-%d"
        ).date()
        invoice.tax_type = request.form.get("tax_type")
        invoice.place_of_supply = request.form.get("place_of_supply")
        invoice.sac_hsn_code = request.form.get("sac_hsn_code")
        invoice.reverse_charge = float(request.form.get("reverse_charge") or 0)
        invoice.is_rcm = request.form.get("is_rcm") == "on"
        invoice.distributor_code = request.form.get("distributor_code", "").strip()

        # Local copy overrides
        invoice.party_name = request.form.get("party_name", "").strip()
        invoice.party_address = request.form.get("party_address", "").strip()
        invoice.party_gstin = request.form.get("party_gstin", "").strip().upper()
        invoice.party_pan = request.form.get("party_pan", "").strip().upper()
        invoice.party_state = request.form.get("party_state", "").strip()
        invoice.party_state_code = request.form.get("party_state_code", "").strip()

        if invoice.is_rcm and invoice.reverse_charge <= 0:
            flash("Reverse Charge amount is required when RCM is enabled.", "danger")
            return redirect(url_for("edit_invoice", invoice_id=invoice_id))

        if invoice.invoice_no and not invoice.total_in_words:
            invoice.total_in_words = number_to_words(invoice.calculate_gst()["total"])

        for item in invoice.items:
            db.session.delete(item)

        items_desc = request.form.getlist("item_description[]")
        items_taxable = request.form.getlist("item_taxable_value[]")
        items_cgst_rate = request.form.getlist("item_cgst_rate[]")
        items_sgst_rate = request.form.getlist("item_sgst_rate[]")
        items_igst_rate = request.form.getlist("item_igst_rate[]")

        tax_type = request.form.get("tax_type")
        is_rcm = request.form.get("is_rcm") == "on"

        for i in range(len(items_desc)):
            if items_desc[i].strip():
                taxable = float(items_taxable[i]) if i < len(items_taxable) else 0
                cgst_rate = float(items_cgst_rate[i]) if i < len(items_cgst_rate) else 0
                sgst_rate = float(items_sgst_rate[i]) if i < len(items_sgst_rate) else 0
                igst_rate = float(items_igst_rate[i]) if i < len(items_igst_rate) else 0

                is_valid, error_msg = validate_tax_rates(
                    tax_type, is_rcm, cgst_rate, sgst_rate, igst_rate
                )
                if not is_valid:
                    flash(f"Item '{items_desc[i][:30]}': {error_msg}", "danger")
                    return redirect(url_for("edit_invoice", invoice_id=invoice_id))

                cgst_amt = taxable * cgst_rate / 100 if tax_type == "INTRA" else 0
                sgst_amt = taxable * sgst_rate / 100 if tax_type == "INTRA" else 0
                igst_amt = taxable * igst_rate / 100 if tax_type == "INTER" else 0

                item = InvoiceItem(
                    invoice_id=invoice.id,
                    description=items_desc[i],
                    taxable_value=taxable,
                    cgst_rate=cgst_rate,
                    cgst_amt=cgst_amt,
                    sgst_rate=sgst_rate,
                    sgst_amt=sgst_amt,
                    igst_rate=igst_rate,
                    igst_amt=igst_amt,
                )
                db.session.add(item)

        db.session.commit()
        flash("Invoice updated successfully", "success")
        return redirect(url_for("manage_invoices"))

    return render_template("edit_invoice.html", invoice=invoice, parties=parties)


def generate_credit_note_numbers():
    current_fy = get_fiscal_year()
    fy_prefix = current_fy.replace("-", "")

    pending_credit_notes = (
        CreditNote.query.filter(CreditNote.credit_note_no == None)
        .order_by(CreditNote.credit_note_date)
        .all()
    )

    if not pending_credit_notes:
        return 0

    last_credit_note = (
        CreditNote.query.filter(CreditNote.credit_note_no != None)
        .filter(CreditNote.credit_note_no.like(f"CRN/{fy_prefix}%"))
        .order_by(CreditNote.credit_note_no.desc())
        .first()
    )

    if last_credit_note:
        last_no = last_credit_note.credit_note_no
        parts = last_no.split("/")
        if len(parts) == 3:
            try:
                seq = int(parts[2])
            except:
                seq = 0
        else:
            seq = 0
    else:
        seq = 0

    for credit_note in pending_credit_notes:
        seq += 1
        credit_note_no = f"CRN/{fy_prefix}/{str(seq).zfill(3)}"
        credit_note.credit_note_no = credit_note_no

        total = credit_note.calculate_gst()["total"]
        credit_note.total_in_words = number_to_words(total)

    db.session.commit()
    return len(pending_credit_notes)


@app.route("/credit-notes")
def manage_credit_notes():
    year = request.args.get("year")
    month = request.args.get("month")
    search = request.args.get("search", "").strip()
    sort_by = request.args.get("sort_by", "credit_note_no")
    sort_dir = request.args.get("sort_dir", "desc")
    selected_year = year
    selected_month = month
    search_query = search
    invoice_search = request.args.get("invoice_search", "").strip()
    status_filter = request.args.get("status", "")

    query = CreditNote.query.join(Invoice).join(Party)

    if year:
        query = query.filter(
            db.extract("year", CreditNote.credit_note_date) == int(year)
        )
    if month:
        query = query.filter(
            db.extract("month", CreditNote.credit_note_date) == int(month)
        )
    if search:
        query = query.filter(
            db.or_(
                CreditNote.credit_note_no.ilike(f"%{search}%"),
                CreditNote.party_name.ilike(f"%{search}%"),
                CreditNote.party_gstin.ilike(f"%{search}%"),
            )
        )
    if invoice_search:
        query = query.filter(
            db.or_(
                Invoice.invoice_no.ilike(f"%{invoice_search}%"),
                Invoice.reference_serial_no.ilike(f"%{invoice_search}%"),
                Party.name.ilike(f"%{invoice_search}%"),
            )
        )
    if status_filter == "pending":
        query = query.filter(CreditNote.credit_note_no == None)
    elif status_filter == "completed":
        query = query.filter(CreditNote.credit_note_no != None)

    sort_column = CreditNote.credit_note_no
    if sort_by == "date":
        sort_column = CreditNote.credit_note_date
    elif sort_by == "invoice":
        sort_column = Invoice.invoice_no
    elif sort_by == "party":
        sort_column = Party.name

    if sort_dir == "desc":
        query = query.order_by(sort_column.desc())
    else:
        query = query.order_by(sort_column.asc())

    credit_notes = query.all()
    pending_count = CreditNote.query.filter(CreditNote.credit_note_no == None).count()

    all_credit_notes = CreditNote.query.all()
    years = sorted(
        set(cn.credit_note_date.year for cn in all_credit_notes if cn.credit_note_date)
    )
    if not years:
        years = [datetime.now().year]

    return render_template(
        "credit_note_management.html",
        credit_notes=credit_notes,
        selected_year=selected_year,
        selected_month=selected_month,
        search_query=search_query,
        years=years,
        sort_by=sort_by,
        sort_dir=sort_dir,
        invoice_search=invoice_search,
        status_filter=status_filter,
        pending_count=pending_count,
    )


@app.route("/credit-note/create", methods=["GET", "POST"])
def create_credit_note():
    existing_invoice_ids = [cn.invoice_id for cn in CreditNote.query.all()]
    if existing_invoice_ids:
        invoices = (
            Invoice.query.filter(Invoice.locked == True)
            .filter(~Invoice.id.in_(existing_invoice_ids))
            .order_by(Invoice.invoice_date.desc())
            .all()
        )
    else:
        invoices = (
            Invoice.query.filter(Invoice.locked == True)
            .order_by(Invoice.invoice_date.desc())
            .all()
        )

    if request.method == "POST":
        invoice_id = request.form.get("invoice_id")
        credit_note_date = datetime.strptime(
            request.form.get("credit_note_date"), "%Y-%m-%d"
        ).date()
        reason = request.form.get("reason")
        tax_type = request.form.get("tax_type")
        place_of_supply = request.form.get("place_of_supply")

        if not invoice_id:
            flash("Invoice selection is required.", "danger")
            return redirect(url_for("create_credit_note"))

        if not reason:
            flash("Reason is required.", "danger")
            return redirect(url_for("create_credit_note"))

        existing_credit_note = CreditNote.query.filter_by(invoice_id=invoice_id).first()
        if existing_credit_note:
            flash(f"Credit note already exists for this invoice.", "warning")
            return redirect(url_for("create_credit_note"))

        invoice = db.session.get(Invoice, invoice_id)

        credit_note = CreditNote(
            credit_note_date=credit_note_date,
            invoice_id=invoice_id,
            reason=reason,
            tax_type=tax_type,
            place_of_supply=place_of_supply,
        )

        credit_note.party_name = invoice.party_name
        credit_note.party_address = invoice.party_address
        credit_note.party_gstin = invoice.party_gstin
        credit_note.party_pan = invoice.party_pan
        credit_note.party_state = invoice.party_state
        credit_note.party_state_code = invoice.party_state_code

        db.session.add(credit_note)
        db.session.flush()

        items_desc = request.form.getlist("item_description[]")
        items_taxable = request.form.getlist("item_taxable_value[]")
        items_cgst_rate = request.form.getlist("item_cgst_rate[]")
        items_sgst_rate = request.form.getlist("item_sgst_rate[]")
        items_igst_rate = request.form.getlist("item_igst_rate[]")

        for i in range(len(items_desc)):
            if items_desc[i].strip():
                taxable = float(items_taxable[i]) if i < len(items_taxable) else 0
                cgst_rate = float(items_cgst_rate[i]) if i < len(items_cgst_rate) else 0
                sgst_rate = float(items_sgst_rate[i]) if i < len(items_sgst_rate) else 0
                igst_rate = float(items_igst_rate[i]) if i < len(items_igst_rate) else 0

                cgst_amt = taxable * cgst_rate / 100 if tax_type == "INTRA" else 0
                sgst_amt = taxable * sgst_rate / 100 if tax_type == "INTRA" else 0
                igst_amt = taxable * igst_rate / 100 if tax_type == "INTER" else 0

                item = CreditNoteItem(
                    credit_note_id=credit_note.id,
                    description=items_desc[i],
                    taxable_value=taxable,
                    cgst_rate=cgst_rate,
                    cgst_amt=cgst_amt,
                    sgst_rate=sgst_rate,
                    sgst_amt=sgst_amt,
                    igst_rate=igst_rate,
                    igst_amt=igst_amt,
                )
                db.session.add(item)

        db.session.commit()
        flash("Credit note created successfully", "success")
        return redirect(url_for("manage_credit_notes"))

    invoice_id = request.args.get("invoice_id")
    selected_invoice = None
    invoice_items = []
    if invoice_id:
        selected_invoice = db.session.get(Invoice, invoice_id)
        if selected_invoice:
            invoice_items = selected_invoice.items

    return render_template(
        "create_credit_note.html",
        invoices=invoices,
        selected_invoice=selected_invoice,
        invoice_items=invoice_items,
    )


@app.route("/credit-note/delete/<int:credit_note_id>")
def delete_credit_note(credit_note_id):
    credit_note = CreditNote.query.get_or_404(credit_note_id)
    if credit_note.credit_note_no:
        flash("Locked credit notes cannot be deleted", "danger")
        return redirect(url_for("manage_credit_notes"))
    db.session.delete(credit_note)
    db.session.commit()
    flash("Credit note deleted successfully", "success")
    return redirect(url_for("manage_credit_notes"))


@app.route("/credit-note/batch-delete", methods=["POST"])
def batch_delete_credit_notes():
    ids_raw = request.form.get("credit_note_ids", "")
    credit_note_ids = ids_raw.split(",") if ids_raw else []
    if not credit_note_ids or credit_note_ids == [""]:
        flash("No credit notes selected", "warning")
        return redirect(url_for("manage_credit_notes"))
    deleted_count = 0
    skipped_count = 0
    for cn_id in credit_note_ids:
        cn = db.session.get(CreditNote, int(cn_id))
        if cn and not cn.locked:
            db.session.delete(cn)
            deleted_count += 1
        else:
            skipped_count += 1
    db.session.commit()
    if deleted_count > 0:
        flash(f"Successfully deleted {deleted_count} credit notes", "success")
    if skipped_count > 0:
        flash(f"Skipped {skipped_count} locked credit note(s)", "warning")
    return redirect(url_for("manage_credit_notes"))


@app.route("/credit-note/batch-lock", methods=["POST"])
def batch_lock_credit_notes():
    ids_raw = request.form.get("credit_note_ids", "")
    credit_note_ids = ids_raw.split(",") if ids_raw else []
    if not credit_note_ids or credit_note_ids == [""]:
        flash("No credit notes selected", "warning")
        return redirect(url_for("manage_credit_notes"))
    locked_count = 0
    skipped_count = 0
    for cn_id in credit_note_ids:
        cn = db.session.get(CreditNote, int(cn_id))
        if cn:
            if cn.locked:
                skipped_count += 1
            elif not cn.credit_note_no:
                skipped_count += 1
            else:
                cn.locked = True
                locked_count += 1
    db.session.commit()
    if locked_count > 0:
        flash(f"Successfully locked {locked_count} credit notes", "success")
    if skipped_count > 0:
        flash(
            f"Skipped {skipped_count} credit note(s) - already locked or without credit note number",
            "info",
        )
    return redirect(url_for("manage_credit_notes"))


@app.route("/credit-note/batch-unlock", methods=["POST"])
def batch_unlock_credit_notes():
    ids_raw = request.form.get("credit_note_ids", "")
    credit_note_ids = ids_raw.split(",") if ids_raw else []
    if not credit_note_ids or credit_note_ids == [""]:
        flash("No credit notes selected", "warning")
        return redirect(url_for("manage_credit_notes"))
    unlocked_count = 0
    skipped_count = 0
    for cn_id in credit_note_ids:
        cn = db.session.get(CreditNote, int(cn_id))
        if cn:
            if not cn.locked:
                skipped_count += 1
            else:
                cn.locked = False
                unlocked_count += 1
    db.session.commit()
    if unlocked_count > 0:
        flash(f"Successfully unlocked {unlocked_count} credit notes", "success")
    if skipped_count > 0:
        flash(f"Skipped {skipped_count} credit note(s) - already unlocked", "info")
    return redirect(url_for("manage_credit_notes"))


@app.route("/credit-note/preview/<int:credit_note_id>")
def preview_credit_note(credit_note_id):
    credit_note = CreditNote.query.get_or_404(credit_note_id)
    return render_template("credit_note_preview.html", credit_note=credit_note)


@app.route("/credit-note/pdf/<int:credit_note_id>")
def generate_credit_note_pdf(credit_note_id):
    credit_note = CreditNote.query.get_or_404(credit_note_id)

    html_content = render_template(
        "credit_note_pdf_template.html",
        credit_note=credit_note,
        settings={
            "company_name": Settings.get("company_name", ""),
            "address": Settings.get("address", ""),
            "gstin": Settings.get("gstin", ""),
            "pan": Settings.get("pan", ""),
            "logo": Settings.get("logo", ""),
            "place_of_supply": Settings.get("place_of_supply", ""),
            "state_code": Settings.get("state_code", ""),
        },
    )

    fy_short = get_fy_short()
    month_name = (
        get_month_name(credit_note.credit_note_date.month)
        if credit_note.credit_note_date
        else "Unknown"
    )
    party_name = sanitize_filename(credit_note.party_name or "Unknown")
    credit_note_no = sanitize_filename(
        credit_note.credit_note_no or f"credit_note_{credit_note.id}"
    )
    pdf_path = os.path.join(
        app.config["EXPORT_FOLDER"],
        fy_short,
        month_name,
        f"CN_{party_name}_{credit_note_no}.pdf",
    )
    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)

    HTML(string=html_content).write_pdf(pdf_path)

    download_name = (
        f"{credit_note.credit_note_no}.pdf"
        if credit_note.credit_note_no
        else f"credit_note_{credit_note_id}.pdf"
    )
    return send_file(pdf_path, as_attachment=True, download_name=download_name)


@app.route("/credit-note/generate-numbers", methods=["POST"])
def generate_credit_note_numbers_route():
    count = generate_credit_note_numbers()
    if count > 0:
        flash(f"Generated credit note numbers for {count} credit notes", "success")
    else:
        flash("No pending credit notes to generate numbers for", "warning")
    return redirect(url_for("manage_credit_notes"))


@app.template_filter("currency")
def currency_filter(value):
    if value is None or value == "":
        return "0.00"
    try:
        return f"{float(value):,.2f}"
    except (ValueError, TypeError):
        return "0.00"


@app.template_filter("max")
def max_filter(value):
    try:
        return max(value) if value else 0
    except (TypeError, ValueError):
        return 0


@app.context_processor
def inject_now():
    from datetime import datetime, timezone

    return dict(
        current_date=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        settings={
            "company_name": Settings.get("company_name", ""),
            "arn_number": Settings.get("arn_number", ""),
            "address": Settings.get("address", ""),
            "gstin": Settings.get("gstin", ""),
            "pan": Settings.get("pan", ""),
            "logo": Settings.get("logo", ""),
            "place_of_supply": Settings.get("place_of_supply", ""),
            "state_code": Settings.get("state_code", ""),
        },
    )


def init_db():
    with app.app_context():
        db.create_all()


if __name__ == "__main__":
    init_db()
    app.run(debug=True, host="0.0.0.0", port=5000)
